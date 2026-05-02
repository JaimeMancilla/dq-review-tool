"""
Cluster Reviewer v4
- Conexión directa a PostgreSQL (credenciales en .env)
- Memoria interna SQLite (correcciones pasadas)
- Subgrupos reagrupables
- Feedback / aprendizaje activo

Correr: python app.py
Abrir:  http://localhost:5000
"""

import re, unicodedata, json, time, sqlite3, os
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, session
import pandas as pd

# Cargar .env si existe
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = "cluster_reviewer_2024"

# Sesiones en filesystem (evita límite de 4KB de cookies)
try:
    from flask_session import Session
    app.config["SESSION_TYPE"]           = "filesystem"
    app.config["SESSION_FILE_DIR"]       = "flask_sessions"
    app.config["SESSION_PERMANENT"]      = False
    app.config["SESSION_USE_SIGNER"]     = True
    Path("flask_sessions").mkdir(exist_ok=True)
    Session(app)
except ImportError:
    pass  # fallback a cookie session

UPLOAD_FOLDER = Path("uploads"); UPLOAD_FOLDER.mkdir(exist_ok=True)
FEEDBACK_FILE = Path("feedback.json")
DB_FILE       = Path("memory.db")
MODEL_NAME = r"C:\models\paraphrase-multilingual"

# ── PostgreSQL ────────────────────────────────────────────────────────────────

def get_pg_config():
    return {
        "host":     os.getenv("PG_HOST",     ""),
        "database": os.getenv("PG_DATABASE", ""),
        "user":     os.getenv("PG_USER",     ""),
        "password": os.getenv("PG_PASSWORD", ""),
        "port":     int(os.getenv("PG_PORT", "5432")),
    }

def pg_is_configured():
    cfg = get_pg_config()
    return bool(cfg["host"] and cfg["database"] and cfg["user"])

def pg_query(sql, timeout_ms=15000):
    """Ejecuta una query y retorna filas como lista de dicts."""
    import psycopg2
    import psycopg2.extras
    cfg = get_pg_config()
    conn = psycopg2.connect(**cfg, connect_timeout=10,
                             options=f"-c statement_timeout={timeout_ms}")
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql)
            rows = [dict(r) for r in cur.fetchall()]
        return rows, None
    except Exception as e:
        return None, str(e)
    finally:
        conn.close()

# ── SQLite memoria interna ────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS corrections (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            item_index    TEXT NOT NULL,
            cluster_id    TEXT,
            app_name      TEXT,
            app_address   TEXT,
            anchor_name   TEXT,
            correction    TEXT,
            is_new        INTEGER DEFAULT 0,
            reviewed_at   REAL,
            file_name     TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_item ON corrections(item_index)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_appname ON corrections(app_name)")
    # Tabla para restauración completa de sesión por archivo
    conn.execute("""
        CREATE TABLE IF NOT EXISTS session_state (
            file_name    TEXT PRIMARY KEY,
            state_json   TEXT NOT NULL,
            saved_at     REAL
        )
    """)
    # Tabla de correcciones externas (stores no en el archivo pero identificados de paso)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS external_corrections (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            store_id         TEXT NOT NULL,
            item_index       TEXT NOT NULL,
            cluster_index    TEXT,
            cluster_name     TEXT,
            cluster_address  TEXT,
            app_name         TEXT,
            app_address      TEXT,
            scraper_source   TEXT,
            correction       TEXT NOT NULL,
            added_at         REAL,
            file_name        TEXT
        )
    """)
    # Migración: agregar scraper_source si no existe
    try:
        conn.execute("ALTER TABLE external_corrections ADD COLUMN scraper_source TEXT")
    except: pass
    # Migración: agregar is_new si no existe (0 = migrada a cluster existente, 1 = generó nuevo cluster)
    try:
        conn.execute("ALTER TABLE external_corrections ADD COLUMN is_new INTEGER DEFAULT 0")
    except: pass
    # Tabla de progreso: clusters revisados
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reviewed_clusters (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            cluster_index   TEXT NOT NULL,
            cluster_name    TEXT,
            had_errors      INTEGER DEFAULT 0,
            corrections_count INTEGER DEFAULT 0,
            reviewed_at     REAL,
            file_name       TEXT,
            review_type     TEXT DEFAULT 'stores_restaurant'
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ci ON reviewed_clusters(cluster_index)")
    # Migración: agregar columna review_type si no existe
    try:
        conn.execute("ALTER TABLE reviewed_clusters ADD COLUMN review_type TEXT DEFAULT 'stores_restaurant'")
    except: pass
    # Pares etiquetados de stores para futuro fine-tuning
    conn.execute("""
        CREATE TABLE IF NOT EXISTS store_pairs (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            app_name_a      TEXT,
            app_address_a   TEXT,
            cluster_index_a TEXT,
            app_name_b      TEXT,
            app_address_b   TEXT,
            cluster_index_b TEXT,
            label           INTEGER,  -- 1=mismo local, 0=distinto
            source          TEXT,     -- "correction" | "manual"
            added_at        REAL,
            file_name       TEXT
        )
    """)
    # Pares etiquetados de dishes para futuro fine-tuning
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dish_pairs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name_a      TEXT, desc_a TEXT,
            name_b      TEXT, desc_b TEXT,
            label       INTEGER,  -- 1=mismo plato, 0=distinto, 2=fusionar
            source      TEXT,     -- "fusion" | "revision"
            added_at    REAL,
            file_name   TEXT
        )
    """)
    # Tabla de archivos revisados (resumen por archivo)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reviewed_files (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name    TEXT NOT NULL,
            review_type  TEXT NOT NULL,
            total        INTEGER DEFAULT 0,
            ok           INTEGER DEFAULT 0,
            bad          INTEGER DEFAULT 0,
            incomplete   INTEGER DEFAULT 0,
            reviewed_at  REAL
        )
    """)
    conn.commit()
    conn.close()

init_db()

# Migración: agregar tabla cluster_scores si no existe
def _migrate_cluster_scores():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cluster_scores (
            cluster_index   TEXT NOT NULL,
            country         TEXT NOT NULL,
            review_type     TEXT NOT NULL DEFAULT 'stores_restaurant',
            cluster_name    TEXT,
            cluster_address TEXT,
            cluster_ciudad  TEXT,
            cluster_estado  TEXT,
            main_chain      TEXT,
            member_count    INTEGER DEFAULT 1,
            score_t1        REAL,
            has_t2          INTEGER DEFAULT 0,
            t2_neighbor     TEXT,
            t2_dist_m       REAL,
            t2_sim          REAL,
            estado_revision TEXT DEFAULT 'pendiente',
            scored_at       REAL,
            PRIMARY KEY (cluster_index, review_type)
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cs_country ON cluster_scores(country, review_type)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cs_ciudad  ON cluster_scores(cluster_ciudad)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cs_chain   ON cluster_scores(main_chain)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cs_estado  ON cluster_scores(estado_revision)")
    conn.commit(); conn.close()

_migrate_cluster_scores()

# ── Migración: tablas para módulo Homologación PJ (Papa John's) ──────────────
def _migrate_pj_tables():
    """Crea las 4 tablas del módulo de Homologación PJ.
    - pj_sabana: filas crudas del Excel del cliente (una por product_id de sabana)
    - pj_master: Maestra de productos (catálogo canónico con search_codes)
    - pj_cross:  Maestra competencia (matriz de equivalencias entre cadenas)
    - pj_homol:  asignaciones (main_chain, product_name) -> search_code [global]
    """
    conn = sqlite3.connect(DB_FILE)
    # Sabana: datos crudos del archivo, se recarga en cada upload del mismo file_name
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pj_sabana (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name           TEXT NOT NULL,
            client              TEXT,
            country             TEXT,
            scraper_source      TEXT,
            main_chain          TEXT,
            store_id            TEXT,
            product_id          TEXT,
            product_name        TEXT,
            product_description TEXT,
            product_category    TEXT,
            price               TEXT,
            url_example         TEXT,
            search_code_original TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjs_file  ON pj_sabana(file_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjs_chain ON pj_sabana(file_name, main_chain)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjs_pname ON pj_sabana(file_name, main_chain, product_name)")

    # Maestra de productos: catálogo del archivo, se recarga en cada upload
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pj_master (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name           TEXT NOT NULL,
            client              TEXT,
            main_chain          TEXT,
            search_code         TEXT,
            search_name         TEXT,
            search_description  TEXT,
            informacion_faltante TEXT DEFAULT ''
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjm_file  ON pj_master(file_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjm_chain ON pj_master(file_name, main_chain)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjm_code  ON pj_master(file_name, search_code)")

    # Maestra competencia: matriz de equivalencias entre cadenas, se recarga en cada upload
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pj_cross (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name       TEXT NOT NULL,
            tipo            TEXT,
            chain_name      TEXT,
            product_name    TEXT,
            row_idx         INTEGER
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjc_file ON pj_cross(file_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjc_row  ON pj_cross(file_name, row_idx)")

    # Homologaciones: persistentes y globales (cross-file) por (main_chain, product_name)
    # Si llega un archivo nuevo, precargamos auto las homologaciones que ya existan.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pj_homol (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            main_chain      TEXT NOT NULL,
            product_name    TEXT NOT NULL,
            search_code     TEXT,
            is_dubious      INTEGER DEFAULT 0,
            note            TEXT DEFAULT '',
            file_name       TEXT,
            updated_at      REAL,
            UNIQUE(main_chain, product_name)
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjh_chain ON pj_homol(main_chain)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pjh_code  ON pj_homol(search_code)")
    conn.commit(); conn.close()

_migrate_pj_tables()

# ── Módulo Homologación PJ: parseo + endpoints ───────────────────────────────
PJ_CURRENT_FILE_KEY = "pj_filename"  # clave de sesión para archivo activo

def _pj_clear_file(conn, file_name):
    """Borra datos de un archivo en pj_sabana, pj_master, pj_cross (no toca pj_homol global)."""
    conn.execute("DELETE FROM pj_sabana WHERE file_name = ?", (file_name,))
    conn.execute("DELETE FROM pj_master WHERE file_name = ?", (file_name,))
    conn.execute("DELETE FROM pj_cross  WHERE file_name = ?", (file_name,))

def _pj_parse_xlsx(path, file_name):
    """Parsea un Excel de homologación PJ. Espera 3 hojas:
    - 'sabana de datos' con 12 columnas conocidas
    - 'Maestra de productos' con 5 columnas
    - 'Maestra competencia' (matriz)
    Retorna (stats_dict, error_str). Si error, no guarda nada."""
    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        return None, f"No se pudo abrir el Excel: {e}"

    sheets = xl.sheet_names
    # Tolerante: buscar hojas por keywords (por si el cliente cambia mayus/min)
    def _find_sheet(keywords):
        for s in sheets:
            sn = s.lower().strip()
            if all(k in sn for k in keywords):
                return s
        return None

    sh_sabana = _find_sheet(["sabana"]) or _find_sheet(["sábana"])
    sh_master = _find_sheet(["maestra", "producto"])
    sh_cross  = _find_sheet(["maestra", "competencia"])

    if not sh_sabana:
        return None, "No se encontró la hoja 'sabana de datos'"
    if not sh_master:
        return None, "No se encontró la hoja 'Maestra de productos'"

    # Parseo de sabana — mantener columnas esperadas, ignorar extras
    sabana = pd.read_excel(path, sheet_name=sh_sabana, dtype=str).fillna("")
    expected_cols = ['client','country','scraper_source','main_chain','store_id',
                     'product_id','product_name','product_description','product_category',
                     'price','url_example','search_code']
    missing = [c for c in expected_cols if c not in sabana.columns]
    if missing:
        return None, f"Faltan columnas en 'sabana de datos': {missing}"

    # Parseo maestra
    master = pd.read_excel(path, sheet_name=sh_master, dtype=str).fillna("")
    expected_m = ['client','main_chain','search_code','search_name','search_description']
    missing_m = [c for c in expected_m if c not in master.columns]
    if missing_m:
        return None, f"Faltan columnas en 'Maestra de productos': {missing_m}"
    # Columna informacion_faltante es opcional en input
    if 'informacion_faltante' not in master.columns:
        master['informacion_faltante'] = ''

    # Parseo cross (opcional). Formato matriz: 1 columna "Tipo" + N columnas de cadenas
    cross_rows = []
    if sh_cross:
        cross = pd.read_excel(path, sheet_name=sh_cross, dtype=str).fillna("")
        if 'Tipo' in cross.columns or 'tipo' in cross.columns:
            tipo_col = 'Tipo' if 'Tipo' in cross.columns else 'tipo'
            chain_cols = [c for c in cross.columns if c != tipo_col]
            for ridx, row in cross.iterrows():
                tipo_v = row.get(tipo_col,"")
                for ch in chain_cols:
                    pname = row.get(ch,"").strip()
                    if pname:
                        cross_rows.append({
                            "tipo": tipo_v, "chain_name": ch.strip(),
                            "product_name": pname, "row_idx": int(ridx)
                        })

    # Todo OK, persistir
    conn = sqlite3.connect(DB_FILE)
    _pj_clear_file(conn, file_name)

    # Sabana
    sabana_rows = sabana[expected_cols].to_dict('records')
    conn.executemany("""
        INSERT INTO pj_sabana (file_name, client, country, scraper_source, main_chain,
                               store_id, product_id, product_name, product_description,
                               product_category, price, url_example, search_code_original)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, [(file_name, r['client'], r['country'], r['scraper_source'], r['main_chain'],
           r['store_id'], r['product_id'], r['product_name'], r['product_description'],
           r['product_category'], r['price'], r['url_example'], r['search_code'])
          for r in sabana_rows])

    # Master
    master_rows = master[expected_m + ['informacion_faltante']].to_dict('records')
    conn.executemany("""
        INSERT INTO pj_master (file_name, client, main_chain, search_code, search_name,
                               search_description, informacion_faltante)
        VALUES (?,?,?,?,?,?,?)
    """, [(file_name, r['client'], r['main_chain'], r['search_code'], r['search_name'],
           r['search_description'], r.get('informacion_faltante',''))
          for r in master_rows])

    # Cross
    if cross_rows:
        conn.executemany("""
            INSERT INTO pj_cross (file_name, tipo, chain_name, product_name, row_idx)
            VALUES (?,?,?,?,?)
        """, [(file_name, r['tipo'], r['chain_name'], r['product_name'], r['row_idx'])
              for r in cross_rows])

    conn.commit(); conn.close()

    return {
        "file_name": file_name,
        "sabana_rows": len(sabana_rows),
        "master_rows": len(master_rows),
        "cross_rows": len(cross_rows),
        "chains_sabana": sorted(sabana['main_chain'].unique().tolist()),
        "chains_master": sorted(master['main_chain'].unique().tolist()),
    }, None

@app.route("/pj/upload", methods=["POST"])
def pj_upload():
    """Recibe un xlsx del cliente PJ, parsea las 3 hojas, guarda en SQLite.
    Mantiene el archivo en disco para permitir export idéntico."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No se recibió archivo"}), 400

    filename = f.filename or "homolog_pj.xlsx"
    # Guardar archivo físico (para poder leer hoja READ_ME y mantener estilos en export)
    pj_dir = os.path.join(os.path.dirname(DB_FILE) or ".", "pj_uploads")
    os.makedirs(pj_dir, exist_ok=True)
    save_path = os.path.join(pj_dir, filename)
    f.save(save_path)

    stats, err = _pj_parse_xlsx(save_path, filename)
    if err:
        return jsonify({"error": err}), 400

    session[PJ_CURRENT_FILE_KEY] = filename
    session["pj_path"] = save_path
    session.modified = True

    # Auto-aplicar homologaciones globales previas a productos de este archivo
    conn = sqlite3.connect(DB_FILE)
    preloaded = conn.execute("""
        SELECT COUNT(DISTINCT s.main_chain || '|' || s.product_name)
        FROM pj_sabana s
        JOIN pj_homol h
          ON h.main_chain = s.main_chain AND h.product_name = s.product_name
        WHERE s.file_name = ? AND h.search_code IS NOT NULL AND h.search_code != ''
    """, (filename,)).fetchone()[0]
    conn.close()
    stats["preloaded_homologations"] = preloaded

    return jsonify({"ok": True, "stats": stats})

@app.route("/pj/session")
def pj_session():
    """Devuelve el archivo PJ activo, si existe."""
    fn = session.get(PJ_CURRENT_FILE_KEY)
    if not fn:
        return jsonify({"active": False})
    conn = sqlite3.connect(DB_FILE)
    n = conn.execute("SELECT COUNT(*) FROM pj_sabana WHERE file_name = ?", (fn,)).fetchone()[0]
    conn.close()
    if n == 0:
        # Archivo borrado del storage, limpiar sesión
        session.pop(PJ_CURRENT_FILE_KEY, None)
        session.pop("pj_path", None)
        session.modified = True
        return jsonify({"active": False})
    return jsonify({"active": True, "file_name": fn, "sabana_rows": n})

@app.route("/pj/list")
def pj_list():
    """Lista agrupada por (main_chain, product_name). Filtros: main_chain, scraper_source,
    status=all|pending|done|dubious, q=búsqueda texto, limit, offset."""
    fn = session.get(PJ_CURRENT_FILE_KEY)
    if not fn:
        return jsonify({"error": "Sin archivo activo"}), 400

    main_chain = (request.args.get("main_chain") or "").strip()
    scraper    = (request.args.get("scraper_source") or "").strip()
    status     = (request.args.get("status") or "all").strip()
    q          = (request.args.get("q") or "").strip().lower()
    limit      = max(1, min(500, int(request.args.get("limit", 100))))
    offset     = max(0, int(request.args.get("offset", 0)))

    where = ["s.file_name = ?"]
    params = [fn]
    if main_chain:
        where.append("s.main_chain = ?")
        params.append(main_chain)
    if scraper:
        where.append("s.scraper_source = ?")
        params.append(scraper)
    if q:
        where.append("LOWER(s.product_name) LIKE ?")
        params.append(f"%{q}%")

    where_sql = " AND ".join(where)

    conn = sqlite3.connect(DB_FILE)
    # Agrupar por (main_chain, product_name). Agregar métricas y left-join con pj_homol
    rows = conn.execute(f"""
        SELECT
          s.main_chain, s.product_name,
          COUNT(*)                    as n_rows,
          MIN(CAST(NULLIF(s.price,'') AS REAL)) as price_min,
          MAX(CAST(NULLIF(s.price,'') AS REAL)) as price_max,
          GROUP_CONCAT(DISTINCT s.scraper_source) as sources,
          h.search_code, h.is_dubious, h.note
        FROM pj_sabana s
        LEFT JOIN pj_homol h
          ON h.main_chain = s.main_chain AND h.product_name = s.product_name
        WHERE {where_sql}
        GROUP BY s.main_chain, s.product_name
    """, params).fetchall()

    # Filtrar por status después del GROUP BY
    if status == "pending":
        rows = [r for r in rows if not r[6] and not r[7]]
    elif status == "done":
        rows = [r for r in rows if r[6]]
    elif status == "dubious":
        rows = [r for r in rows if r[7]]

    total = len(rows)
    # Ordenar por n_rows desc (más impactantes primero)
    rows.sort(key=lambda r: -r[2])
    page = rows[offset:offset+limit]

    # Total global (sin filtros) para el header
    global_stats = conn.execute("""
        SELECT
          COUNT(DISTINCT s.main_chain || '||' || s.product_name) as total_unique,
          COUNT(*) as total_rows,
          (SELECT COUNT(DISTINCT h2.main_chain || '||' || h2.product_name)
             FROM pj_sabana s2
             JOIN pj_homol h2
               ON h2.main_chain = s2.main_chain AND h2.product_name = s2.product_name
            WHERE s2.file_name = ?
              AND h2.search_code IS NOT NULL AND h2.search_code != ''
          ) as total_done,
          (SELECT COUNT(DISTINCT h3.main_chain || '||' || h3.product_name)
             FROM pj_sabana s3
             JOIN pj_homol h3
               ON h3.main_chain = s3.main_chain AND h3.product_name = s3.product_name
            WHERE s3.file_name = ?
              AND h3.is_dubious = 1
          ) as total_dubious
        FROM pj_sabana s
        WHERE s.file_name = ?
    """, (fn, fn, fn)).fetchone()

    # Opciones para filtros
    chains = [r[0] for r in conn.execute("""
        SELECT DISTINCT main_chain FROM pj_sabana WHERE file_name = ? ORDER BY main_chain
    """, (fn,)).fetchall()]
    scrapers = [r[0] for r in conn.execute("""
        SELECT DISTINCT scraper_source FROM pj_sabana WHERE file_name = ? ORDER BY scraper_source
    """, (fn,)).fetchall()]
    chains_with_master = [r[0] for r in conn.execute("""
        SELECT DISTINCT main_chain FROM pj_master WHERE file_name = ? ORDER BY main_chain
    """, (fn,)).fetchall()]
    conn.close()

    return jsonify({
        "items": [{
            "main_chain": r[0], "product_name": r[1],
            "n_rows": r[2], "price_min": r[3], "price_max": r[4],
            "sources": (r[5] or "").split(","),
            "search_code": r[6] or "",
            "is_dubious": bool(r[7]),
            "note": r[8] or "",
            "has_master": r[0] in chains_with_master,
        } for r in page],
        "total_filtered": total,
        "offset": offset,
        "limit": limit,
        "stats": {
            "total_unique": global_stats[0] or 0,
            "total_rows":   global_stats[1] or 0,
            "total_done":   global_stats[2] or 0,
            "total_dubious":global_stats[3] or 0,
        },
        "filters": {
            "chains": chains,
            "scrapers": scrapers,
            "chains_with_master": chains_with_master,
        }
    })

@app.route("/pj/product_detail")
def pj_product_detail():
    """Detalle de un (main_chain, product_name): descripciones únicas, precios, URLs."""
    fn = session.get(PJ_CURRENT_FILE_KEY)
    if not fn:
        return jsonify({"error": "Sin archivo activo"}), 400
    mc = (request.args.get("main_chain") or "").strip()
    pn = (request.args.get("product_name") or "").strip()
    if not mc or not pn:
        return jsonify({"error": "Faltan parámetros"}), 400

    conn = sqlite3.connect(DB_FILE)
    # Todas las filas del producto en sabana
    rows = conn.execute("""
        SELECT scraper_source, store_id, product_id, product_description, product_category,
               price, url_example
        FROM pj_sabana
        WHERE file_name = ? AND main_chain = ? AND product_name = ?
    """, (fn, mc, pn)).fetchall()

    # Descripciones únicas (no vacías)
    descs = {}
    for r in rows:
        d = (r[3] or "").strip()
        if d:
            descs[d] = descs.get(d, 0) + 1
    descs_sorted = sorted(descs.items(), key=lambda x: -x[1])[:10]

    # Categorías únicas
    cats = sorted({(r[4] or "").strip() for r in rows if (r[4] or "").strip()})
    # Sources
    sources = sorted({(r[0] or "").strip() for r in rows if (r[0] or "").strip()})
    # URLs de ejemplo (una por scraper_source, hasta 3)
    url_examples = {}
    for r in rows:
        src = r[0]; u = r[6]
        if src and u and src not in url_examples:
            url_examples[src] = u
        if len(url_examples) >= 3:
            break

    # Precios
    prices = [float(r[5]) for r in rows if r[5] not in (None,"") and str(r[5]).replace('.','').isdigit()]
    price_min = min(prices) if prices else None
    price_max = max(prices) if prices else None
    price_avg = (sum(prices)/len(prices)) if prices else None

    # Homologación existente
    h = conn.execute("""
        SELECT search_code, is_dubious, note FROM pj_homol
        WHERE main_chain = ? AND product_name = ?
    """, (mc, pn)).fetchone()

    # Maestra del main_chain (para sugerir y para mostrar en panel derecho)
    master = conn.execute("""
        SELECT search_code, search_name, search_description, informacion_faltante
        FROM pj_master
        WHERE file_name = ? AND main_chain = ?
        ORDER BY search_code
    """, (fn, mc)).fetchall()

    # Sugerencia básica: match substring case-insensitive de search_name con product_name
    pn_lower = pn.lower()
    suggestions = []
    for m in master:
        sc, sn, sd, _info = m
        if not sn: continue
        sn_lower = sn.lower()
        if sn_lower == pn_lower:
            suggestions.append({"search_code": sc, "search_name": sn, "match_type": "exacto", "score": 1.0})
        elif sn_lower in pn_lower or pn_lower in sn_lower:
            suggestions.append({"search_code": sc, "search_name": sn, "match_type": "substring", "score": 0.7})
    suggestions = sorted(suggestions, key=lambda x: -x["score"])[:5]

    conn.close()

    return jsonify({
        "main_chain": mc,
        "product_name": pn,
        "n_rows": len(rows),
        "descriptions": [{"text": d, "count": c} for d, c in descs_sorted],
        "categories": cats,
        "sources": sources,
        "url_examples": url_examples,
        "price": {"min": price_min, "max": price_max, "avg": price_avg},
        "homologation": {
            "search_code": h[0] if h else "",
            "is_dubious": bool(h[1]) if h else False,
            "note": h[2] if h else "",
        },
        "master": [{
            "search_code": m[0], "search_name": m[1],
            "search_description": m[2], "informacion_faltante": m[3] or ""
        } for m in master],
        "suggestions": suggestions,
    })

@app.route("/pj/homologate", methods=["POST"])
def pj_homologate():
    """Asigna search_code a un (main_chain, product_name). Persiste global."""
    data = request.json or {}
    mc = (data.get("main_chain") or "").strip()
    pn = (data.get("product_name") or "").strip()
    sc = (data.get("search_code") or "").strip()
    if not mc or not pn:
        return jsonify({"error": "Faltan main_chain / product_name"}), 400
    fn = session.get(PJ_CURRENT_FILE_KEY) or ""

    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO pj_homol (main_chain, product_name, search_code, file_name, updated_at)
        VALUES (?,?,?,?,?)
        ON CONFLICT(main_chain, product_name) DO UPDATE SET
            search_code = excluded.search_code,
            file_name = excluded.file_name,
            updated_at = excluded.updated_at
    """, (mc, pn, sc, fn, time.time()))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/pj/mark_dubious", methods=["POST"])
def pj_mark_dubious():
    """Toggle de is_dubious (marca amarilla). Si no existe el registro, lo crea."""
    data = request.json or {}
    mc = (data.get("main_chain") or "").strip()
    pn = (data.get("product_name") or "").strip()
    val = 1 if data.get("is_dubious") else 0
    if not mc or not pn:
        return jsonify({"error": "Faltan main_chain / product_name"}), 400
    fn = session.get(PJ_CURRENT_FILE_KEY) or ""

    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO pj_homol (main_chain, product_name, is_dubious, file_name, updated_at)
        VALUES (?,?,?,?,?)
        ON CONFLICT(main_chain, product_name) DO UPDATE SET
            is_dubious = excluded.is_dubious,
            updated_at = excluded.updated_at
    """, (mc, pn, val, fn, time.time()))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/pj/note", methods=["POST"])
def pj_set_note():
    """Nota libre del revisor sobre el product_name. Distinto de informacion_faltante
    de la maestra (que es por search_code)."""
    data = request.json or {}
    mc = (data.get("main_chain") or "").strip()
    pn = (data.get("product_name") or "").strip()
    note = (data.get("note") or "").strip()
    if not mc or not pn:
        return jsonify({"error": "Faltan main_chain / product_name"}), 400
    fn = session.get(PJ_CURRENT_FILE_KEY) or ""

    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO pj_homol (main_chain, product_name, note, file_name, updated_at)
        VALUES (?,?,?,?,?)
        ON CONFLICT(main_chain, product_name) DO UPDATE SET
            note = excluded.note,
            updated_at = excluded.updated_at
    """, (mc, pn, note, fn, time.time()))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/pj/master_info", methods=["POST"])
def pj_master_info():
    """Actualiza informacion_faltante de un search_code de la maestra del archivo actual."""
    data = request.json or {}
    sc = (data.get("search_code") or "").strip()
    info = (data.get("informacion_faltante") or "").strip()
    fn = session.get(PJ_CURRENT_FILE_KEY) or ""
    if not sc or not fn:
        return jsonify({"error": "Falta search_code o archivo activo"}), 400
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        UPDATE pj_master SET informacion_faltante = ?
        WHERE file_name = ? AND search_code = ?
    """, (info, fn, sc))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@app.route("/pj/export")
def pj_export():
    """Exporta un xlsx manteniendo estructura del original. Rellena search_code en sabana
    y agrega columna informacion_faltante en la maestra."""
    fn = session.get(PJ_CURRENT_FILE_KEY)
    path = session.get("pj_path")
    if not fn or not path or not os.path.exists(path):
        return jsonify({"error": "Sin archivo activo"}), 400

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    import io

    wb = load_workbook(path)

    # Construir lookup de homologaciones para este archivo
    conn = sqlite3.connect(DB_FILE)
    homol_rows = conn.execute("""
        SELECT DISTINCT s.main_chain, s.product_name, h.search_code, h.is_dubious
        FROM pj_sabana s
        JOIN pj_homol h
          ON h.main_chain = s.main_chain AND h.product_name = s.product_name
        WHERE s.file_name = ?
    """, (fn,)).fetchall()
    homol_map = {(r[0], r[1]): (r[2] or "", bool(r[3])) for r in homol_rows}
    # Map de informacion_faltante por search_code
    info_rows = conn.execute("""
        SELECT search_code, informacion_faltante FROM pj_master WHERE file_name = ?
    """, (fn,)).fetchall()
    info_map = {r[0]: (r[1] or "") for r in info_rows}
    conn.close()

    yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # 1) Actualizar hoja "sabana de datos"
    sh_sabana = None
    for s in wb.sheetnames:
        if "sabana" in s.lower() or "sábana" in s.lower():
            sh_sabana = wb[s]; break
    if sh_sabana:
        headers = [c.value for c in sh_sabana[1]]
        try:
            col_mc = headers.index("main_chain") + 1
            col_pn = headers.index("product_name") + 1
            col_sc = headers.index("search_code") + 1
        except ValueError:
            return jsonify({"error": "Estructura de sabana inesperada"}), 400

        for row_idx in range(2, sh_sabana.max_row + 1):
            mc = sh_sabana.cell(row=row_idx, column=col_mc).value or ""
            pn = sh_sabana.cell(row=row_idx, column=col_pn).value or ""
            key = (str(mc), str(pn))
            if key in homol_map:
                sc, dub = homol_map[key]
                if sc:
                    cell = sh_sabana.cell(row=row_idx, column=col_sc, value=sc)
                    if dub:
                        cell.fill = yellow

    # 2) Actualizar hoja "Maestra de productos" — agregar columna informacion_faltante si no existe
    sh_master = None
    for s in wb.sheetnames:
        sl = s.lower()
        if "maestra" in sl and "producto" in sl:
            sh_master = wb[s]; break
    if sh_master:
        headers = [c.value for c in sh_master[1]]
        col_info = None
        if "informacion_faltante" in headers:
            col_info = headers.index("informacion_faltante") + 1
        else:
            col_info = sh_master.max_column + 1
            sh_master.cell(row=1, column=col_info, value="informacion_faltante")
        col_sc_m = headers.index("search_code") + 1 if "search_code" in headers else None
        if col_sc_m:
            for row_idx in range(2, sh_master.max_row + 1):
                sc = sh_master.cell(row=row_idx, column=col_sc_m).value or ""
                if sc in info_map and info_map[sc]:
                    sh_master.cell(row=row_idx, column=col_info, value=info_map[sc])

    # Enviar como download
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    from flask import send_file
    # Preservar nombre del archivo original con sufijo _homologado
    base, ext = os.path.splitext(fn)
    out_name = f"{base}_homologado{ext or '.xlsx'}"
    return send_file(out, as_attachment=True, download_name=out_name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/pj/reset", methods=["POST"])
def pj_reset():
    """Borra datos del archivo PJ activo de pj_sabana, pj_master, pj_cross.
    NO toca pj_homol (son homologaciones globales persistentes)."""
    fn = session.get(PJ_CURRENT_FILE_KEY)
    if not fn:
        return jsonify({"error": "Sin archivo activo"}), 400
    conn = sqlite3.connect(DB_FILE)
    _pj_clear_file(conn, fn)
    conn.commit(); conn.close()
    # Borrar archivo físico
    p = session.get("pj_path")
    if p and os.path.exists(p):
        try: os.remove(p)
        except: pass
    session.pop(PJ_CURRENT_FILE_KEY, None)
    session.pop("pj_path", None)
    session.modified = True
    return jsonify({"ok": True})

def run_scoring_job(country, city=None, review_type="stores_restaurant",
                    dist_m=50, batch_size=2000):
    """
    Calcula score_t1 (similitud miembro vs ancla, Jaccard) y score_t2
    (clusters cercanos con nombre similar) para un país/ciudad.
    Guarda resultados en cluster_scores (SQLite). Diseñado para background thread.
    """
    import psycopg2, psycopg2.extras
    from collections import defaultdict
    cfg = get_pg_config()
    if not cfg: return {"error": "Sin config PG"}

    table = ("sales_opportunity.dim_maestra_retail"
             if review_type == "stores_retail"
             else "sales_opportunity.dim_maestra")

    city_filter = "AND cluster_ciudad ILIKE %s" if city else ""
    city_param  = [f"%{city}%"] if city else []

    try:
        pg  = psycopg2.connect(**cfg, connect_timeout=30)
        cur = pg.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        # ── 1. Traer anchors ─────────────────────────────────────────────────
        q_anchors = f"""
            SELECT DISTINCT ON (cluster_index)
                cluster_index, cluster_name, cluster_address,
                cluster_ciudad, cluster_estado, main_chain,
                cluster_latitude, cluster_longitude,
                COUNT(*) OVER (PARTITION BY cluster_index) AS member_count
            FROM {table}
            WHERE country = %s {city_filter}
            ORDER BY cluster_index
        """
        cur.execute(q_anchors, [country] + city_param)
        anchors = [dict(r) for r in cur.fetchall()]
        print(f"[scoring] {len(anchors)} anchors para {country}/{city or 'all'}")

        # ── 2. Score T1: similitud mínima miembro vs ancla (Jaccard) ────────
        t1_scores = {}
        for i in range(0, len(anchors), batch_size):
            batch = anchors[i:i+batch_size]
            cids  = [a["cluster_index"] for a in batch]
            ph    = ",".join(["%s"] * len(cids))
            cur.execute(f"""
                SELECT cluster_index, app_name,
                       (item_index = cluster_index) AS is_anchor
                FROM {table}
                WHERE cluster_index IN ({ph})
                  AND app_name IS NOT NULL AND app_name != ''
            """, cids)
            rows = [dict(r) for r in cur.fetchall()]
            by_cluster = defaultdict(list)
            for r in rows:
                by_cluster[r["cluster_index"]].append(r)
            anchor_map = {a["cluster_index"]: a for a in batch}
            for cid, members in by_cluster.items():
                anchor_name = anchor_map.get(cid, {}).get("cluster_name") or ""
                non_anchors = [m for m in members if not m["is_anchor"]]
                if not non_anchors:
                    t1_scores[cid] = 1.0
                    continue
                sims = [jaccard_sim(anchor_name, m["app_name"] or "") for m in non_anchors]
                t1_scores[cid] = round(min(sims), 3)
            print(f"[scoring] T1 lote {i//batch_size+1}: {len(by_cluster)} clusters")

        # ── 3. Score T2: clusters cercanos con nombre similar ────────────────
        cur.execute(f"""
            WITH anch AS (
                SELECT DISTINCT ON (cluster_index)
                    cluster_index, cluster_name,
                    cluster_latitude, cluster_longitude
                FROM {table}
                WHERE country = %s {city_filter}
                  AND cluster_latitude IS NOT NULL
                ORDER BY cluster_index
            )
            SELECT
                a.cluster_index AS ci_a, a.cluster_name AS cn_a,
                b.cluster_index AS ci_b, b.cluster_name AS cn_b,
                111320 * sqrt(
                    power(a.cluster_latitude  - b.cluster_latitude, 2) +
                    power((a.cluster_longitude - b.cluster_longitude)
                          * cos(radians(a.cluster_latitude)), 2)
                ) AS dist_m
            FROM anch a
            JOIN anch b ON a.cluster_index < b.cluster_index
            WHERE 111320 * sqrt(
                    power(a.cluster_latitude  - b.cluster_latitude, 2) +
                    power((a.cluster_longitude - b.cluster_longitude)
                          * cos(radians(a.cluster_latitude)), 2)
                  ) < %s
            ORDER BY dist_m
        """, [country] + city_param + [dist_m])
        t2_rows = [dict(r) for r in cur.fetchall()]
        print(f"[scoring] T2 pares cercanos: {len(t2_rows)}")
        pg.close()

        t2_map = {}
        for r in t2_rows:
            sim = jaccard_sim(r["cn_a"] or "", r["cn_b"] or "")
            for ci, neighbor in [(r["ci_a"], r["ci_b"]), (r["ci_b"], r["ci_a"])]:
                ex = t2_map.get(ci)
                if ex is None or sim > ex["t2_sim"]:
                    t2_map[ci] = {
                        "t2_neighbor": neighbor,
                        "t2_dist_m":   round(r["dist_m"], 1),
                        "t2_sim":      round(sim, 3),
                        "has_t2":      1 if sim >= 0.5 else 0
                    }

        # ── 4. Guardar en SQLite ─────────────────────────────────────────────
        now = time.time()
        sc  = sqlite3.connect(DB_FILE)
        for anchor in anchors:
            cid = anchor["cluster_index"]
            t1  = t1_scores.get(cid, 1.0)
            t2  = t2_map.get(cid, {"has_t2": 0, "t2_neighbor": None,
                                    "t2_dist_m": None, "t2_sim": None})
            sc.execute("""
                INSERT INTO cluster_scores
                    (cluster_index, country, review_type, cluster_name,
                     cluster_address, cluster_ciudad, cluster_estado,
                     main_chain, member_count, score_t1, has_t2,
                     t2_neighbor, t2_dist_m, t2_sim, estado_revision, scored_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                    COALESCE((SELECT estado_revision FROM cluster_scores
                               WHERE cluster_index=? AND review_type=?),
                             'pendiente'), ?)
                ON CONFLICT(cluster_index, review_type) DO UPDATE SET
                    cluster_name=excluded.cluster_name,
                    cluster_address=excluded.cluster_address,
                    cluster_ciudad=excluded.cluster_ciudad,
                    cluster_estado=excluded.cluster_estado,
                    main_chain=excluded.main_chain,
                    member_count=excluded.member_count,
                    score_t1=excluded.score_t1,
                    has_t2=excluded.has_t2,
                    t2_neighbor=excluded.t2_neighbor,
                    t2_dist_m=excluded.t2_dist_m,
                    t2_sim=excluded.t2_sim,
                    scored_at=excluded.scored_at
            """, (
                cid, country, review_type,
                anchor.get("cluster_name"), anchor.get("cluster_address"),
                anchor.get("cluster_ciudad"), anchor.get("cluster_estado"),
                anchor.get("main_chain"), anchor.get("member_count", 1),
                t1, t2["has_t2"], t2.get("t2_neighbor"),
                t2.get("t2_dist_m"), t2.get("t2_sim"),
                cid, review_type, now
            ))
        sc.commit(); sc.close()
        saved = len(anchors)
        print(f"[scoring] ✓ {saved} clusters guardados")
        return {"ok": True, "clusters": saved, "t2_pairs": len(t2_rows)}

    except Exception as e:
        import traceback
        print(f"[scoring] ERROR: {e}\n{traceback.format_exc()}")
        return {"error": str(e)}

def mem_save(records):
    """Guarda correcciones en la memoria interna."""
    conn = sqlite3.connect(DB_FILE)
    for r in records:
        conn.execute("""
            INSERT INTO corrections
              (item_index, cluster_id, app_name, app_address, anchor_name,
               correction, is_new, reviewed_at, file_name)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            r.get("item_index",""), r.get("cluster_id",""),
            r.get("app_name",""),   r.get("app_address",""),
            r.get("anchor_name",""),r.get("correction",""),
            1 if r.get("is_new") else 0,
            time.time(), r.get("file_name","")
        ))
    conn.commit(); conn.close()

def mem_lookup_item(item_index):
    """Busca un item_index exacto en memoria."""
    conn = sqlite3.connect(DB_FILE)
    row = conn.execute(
        "SELECT * FROM corrections WHERE item_index=? ORDER BY reviewed_at DESC LIMIT 1",
        (item_index,)
    ).fetchone()
    conn.close()
    if row:
        cols = ["id","item_index","cluster_id","app_name","app_address",
                "anchor_name","correction","is_new","reviewed_at","file_name"]
        return dict(zip(cols, row))
    return None

def mem_lookup_similar(app_name, anchor_name, threshold=0.75):
    """Busca correcciones previas con nombre similar."""
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT * FROM corrections WHERE correction != '' ORDER BY reviewed_at DESC LIMIT 500"
    ).fetchall()
    conn.close()
    cols = ["id","item_index","cluster_id","app_name","app_address",
            "anchor_name","correction","is_new","reviewed_at","file_name"]
    best, best_sim = None, 0
    for row in rows:
        r = dict(zip(cols, row))
        sim_name   = jaccard_sim(r["app_name"],   app_name)
        sim_anchor = jaccard_sim(r["anchor_name"], anchor_name)
        combined   = sim_name * 0.7 + sim_anchor * 0.3
        if combined > best_sim and combined >= threshold:
            best_sim = combined; best = r
    return best, best_sim

def save_session_state(filename, clusters, subgroups=None):
    """Guarda el estado completo: revisiones, correcciones y subgrupos."""
    state = []
    for cl in clusters:
        state.append({
            "cluster_id":  cl["cluster_id"],
            "corrections": cl.get("corrections", {}),
            "subgroups":   subgroups.get(cl["cluster_id"], []) if subgroups else [],
            "members":     [
                {"item_index": m["item_index"], "revision": m["revision"]}
                for m in cl["members"] if not m["is_anchor"]
            ]
        })
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT OR REPLACE INTO session_state (file_name, state_json, saved_at)
        VALUES (?, ?, ?)
    """, (filename, json.dumps(state, ensure_ascii=False), time.time()))
    conn.commit(); conn.close()

def load_session_state(filename):
    """Restaura el estado guardado para un archivo."""
    conn = sqlite3.connect(DB_FILE)
    row = conn.execute(
        "SELECT state_json, saved_at FROM session_state WHERE file_name=?",
        (filename,)
    ).fetchone()
    conn.close()
    if row:
        return json.loads(row[0]), row[1]
    return None, None

def apply_session_state(clusters, state):
    """Aplica el estado guardado sobre los clusters recién cargados."""
    state_map = {s["cluster_id"]: s for s in state}
    subgroups_restored = {}
    for cl in clusters:
        saved = state_map.get(cl["cluster_id"])
        if not saved: continue
        rev_map = {m["item_index"]: m["revision"] for m in saved["members"]}
        for m in cl["members"]:
            if not m["is_anchor"] and m["item_index"] in rev_map:
                m["revision"] = rev_map[m["item_index"]]
        cl["corrections"] = saved.get("corrections", {})
        cl["ok_count"]  = sum(1 for m in cl["members"] if m["revision"]==1)
        cl["bad_count"] = sum(1 for m in cl["members"] if m["revision"]==0)
        if saved.get("subgroups"):
            subgroups_restored[cl["cluster_id"]] = saved["subgroups"]
    return clusters, subgroups_restored

def mem_stats():
    conn = sqlite3.connect(DB_FILE)
    total   = conn.execute("SELECT COUNT(*) FROM corrections").fetchone()[0]
    unique  = conn.execute("SELECT COUNT(DISTINCT item_index) FROM corrections").fetchone()[0]
    recent  = conn.execute(
        "SELECT item_index, app_name, correction, reviewed_at FROM corrections ORDER BY reviewed_at DESC LIMIT 5"
    ).fetchall()
    conn.close()
    return {"total": total, "unique_items": unique, "recent": recent}

def mem_search(query, limit=50):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute("""
        SELECT DISTINCT item_index, app_name, app_address, correction, reviewed_at
        FROM corrections
        WHERE app_name LIKE ? OR item_index LIKE ? OR correction LIKE ?
        ORDER BY reviewed_at DESC LIMIT ?
    """, (f"%{query}%", f"%{query}%", f"%{query}%", limit)).fetchall()
    conn.close()
    cols = ["item_index","app_name","app_address","correction","reviewed_at"]
    return [dict(zip(cols, r)) for r in rows]

# ── modelo semántico ──────────────────────────────────────────────────────────

_model = None; _model_ready = False; _model_error = None

def get_model():
    global _model, _model_ready, _model_error
    if _model_ready or _model_error: return _model
    try:
        from sentence_transformers import SentenceTransformer
        import os
        # Forzar uso de caché sin contactar HuggingFace
        os.environ["HF_HUB_OFFLINE"] = "1"
        try:
            _model = SentenceTransformer(MODEL_NAME)
        finally:
            # Restaurar para no afectar otras operaciones
            os.environ.pop("HF_HUB_OFFLINE", None)
        _model_ready = True
    except Exception as e:
        _model_error = str(e)
    return _model

def batch_semantic_sim(anchor_name, anchor_addr, members):
    m = get_model()
    if m is None:
        return [jaccard_sim(mem.get("app_name",""), anchor_name)*0.85 +
                jaccard_sim(mem.get("app_address",""), anchor_addr)*0.15
                for mem in members]
    from sentence_transformers import util
    ea_n = m.encode(anchor_name, convert_to_tensor=True, show_progress_bar=False)
    ea_a = m.encode(anchor_addr, convert_to_tensor=True, show_progress_bar=False)
    names = [mem.get("app_name","") for mem in members]
    addrs = [mem.get("app_address","") for mem in members]
    en = m.encode(names, convert_to_tensor=True, show_progress_bar=False)
    ea = m.encode(addrs, convert_to_tensor=True, show_progress_bar=False)
    return [round(float(util.cos_sim(ea_n,en[i]))*0.85 + float(util.cos_sim(ea_a,ea[i]))*0.15, 3)
            for i in range(len(members))]

# ── helpers ───────────────────────────────────────────────────────────────────

def norm(s):
    if not isinstance(s, str): return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode()
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", s.lower())).strip()

def jaccard_sim(a, b):
    wa = set(w for w in norm(a).split() if len(w) > 2)
    wb = set(w for w in norm(b).split() if len(w) > 2)
    if not wa or not wb: return 0.0
    return len(wa & wb) / len(wa | wb)

def load_file(path):
    path = str(path)
    if path.endswith(".csv"):
        for enc in ["utf-8","latin-1"]:
            try: df = pd.read_csv(path, dtype=str, encoding=enc); break
            except: continue
    else:
        df = pd.read_excel(path, dtype=str)
    df = df.fillna(""); df.columns = [c.strip() for c in df.columns]
    return df

def effective_cluster(row):
    nc = row.get("new_cluster","")
    if isinstance(nc,str) and nc.strip(): return nc.strip()
    ci = row.get("cluster_index","")
    return ci.strip() if isinstance(ci,str) else ""

def detect_country(df):
    if "country" in df.columns:
        vals = df["country"].dropna().str.strip().str.lower()
        vals = vals[vals != ""]
        if len(vals): return vals.mode()[0]
    return "mx"

def get_threshold():
    fb = load_feedback()
    return round(min(1.0, max(0.5, 0.95 + fb.get("threshold_adj", 0.0))), 4)

# ── feedback ──────────────────────────────────────────────────────────────────

def load_feedback():
    if FEEDBACK_FILE.exists():
        try: return json.loads(FEEDBACK_FILE.read_text())
        except: pass
    return {"pairs":[], "threshold_adj":0.0, "stats":{"total":0,"overrides_to_1":0,"overrides_to_0":0}}

def save_feedback(fb): FEEDBACK_FILE.write_text(json.dumps(fb, ensure_ascii=False, indent=2))

def record_feedback(anchor_name, anchor_addr, member_name, member_addr, model_score, model_rev, human_rev):
    if model_rev == human_rev: return
    fb = load_feedback()
    fb["pairs"].append({"anchor_name":anchor_name,"anchor_addr":anchor_addr,
                         "member_name":member_name,"member_addr":member_addr,
                         "model_score":model_score,"model_rev":model_rev,
                         "human_rev":human_rev,"ts":time.time()})
    fb["stats"]["total"] += 1
    if human_rev==1: fb["stats"]["overrides_to_1"] += 1
    else:            fb["stats"]["overrides_to_0"] += 1
    recent = fb["pairs"][-30:]
    if len(recent) >= 5:
        fp = sum(1 for p in recent if p["model_rev"]==1 and p["human_rev"]==0)
        fn = sum(1 for p in recent if p["model_rev"]==0 and p["human_rev"]==1)
        adj = max(-0.10, min(0.10, (fp-fn)/len(recent)*0.10))
        fb["threshold_adj"] = round(adj, 4)
    save_feedback(fb)

# ── subgrupos ─────────────────────────────────────────────────────────────────

GEO_STOP = {
    # Colonias / barrios MX frecuentes
    "reforma","serdan","serdán","canek","manuel","moreno","loza","morelos",
    "monica","mónica","hidalgo","centro","pedregal","polanco","roma","condesa",
    "satelite","satélite","narvarte","coyoacan","coyoacán","xochimilco",
    "tepito","tlalpan","iztapalapa","ecatepec","naucalpan","tlalnepantla",
    "texcoco","chalco","ixtapaluca","nezahualcoyotl","nezahualcóyotl",
    "insurgentes","chapultepec","juarez","juárez","guerrero","doctores",
    "lindavista","vallejo","azcapotzalco","magdalena","contreras",
    "tlahuac","tláhuac","milpa","alta","alvaro","obregon","obregón",
    "benito","cuauhtemoc","cuauhtémoc","miguel","venustiano","carranza",
    "iztacalco","gustavo","madero","cuajimalpa","cuautitlan","cuautitlán",
    # Palabras geográficas genéricas
    "san","santa","santo","nuevo","nueva","gran","grande","real","real",
    "jardines","residencial","fracc","fraccionamiento","unidad","barrio",
    "colonia","col","pueblo","ciudad","municipio","delegacion","delegación",
    "local","plaza","paseo","calle","avenida","blvd","boulevard","calz","calzada",
    "sur","norte","oriente","poniente","este","oeste","inner","outer",
    "zona","sector","modulo","módulo","piso","nivel","torre","edificio",
    # Ciudades/estados MX
    "puebla","monterrey","guadalajara","tijuana","juarez","merida","mérida",
    "leon","león","toluca","queretaro","querétaro","chihuahua","hermosillo",
    "mexicali","aguascalientes","morelia","veracruz","oaxaca","tuxtla",
    "gutierrez","gutiérrez","villahermosa","tampico","culiacan","culiacán",
    "durango","zacatecas","saltillo","tepic","colima","campeche","chetumal",
    "pachuca","tlaxcala","cuernavaca","chilpancingo","xalapa","jalapa",
    "cdmx","cdmex","edomex","mx","mexico","méxico",
    # Colombia
    "bogota","bogotá","medellin","medellín","cali","barranquilla","cartagena",
    "cucuta","cúcuta","bucaramanga","pereira","manizales","armenia","ibague",
    "ibagué","neiva","villavicencio","pasto","monteria","montería","sincelejo",
    "valledupar","santa","marta","riohacha","quibdo","quibdó","popayan",
    "popayán","tunja","manizales","florencia","mocoa","leticia","inirida",
    "inírida","mitu","mitú","yopal","arauca","puerto","carreño","san","andres",
    # Brasil
    "sao","são","paulo","rio","janeiro","belo","horizonte","brasilia","brasília",
    "salvador","fortaleza","curitiba","manaus","recife","porto","alegre",
    "belem","belém","goiania","goiânia","florianopolis","florianópolis",
    "natal","maceio","maceió","joao","joão","pessoa","teresina","campo",
    "grande","aracaju","macapa","macapá","porto","velho","palmas","boa","vista",
    "rio","branco","vitoria","vitória","cuiaba","cuiabá",
}

def brand_words(name):
    """
    Extrae palabras de marca ignorando sufijos geográficos.
    Para cadenas con guión (ej: "Castaño - Rosario Norte"), usa solo la parte
    antes del guión para evitar incluir nombres de locales como palabras de marca.
    """
    # Cortar en el primer guión o coma — separa cadena de nombre de local
    if ' - ' in name:
        name = name.split(' - ')[0].strip()
    elif '-' in name:
        name = name.split('-')[0].strip()
    if ',' in name:
        name = name.split(',')[0].strip()
    func = {"la","el","los","las","de","del","en","al","por","con","que","y","e","o"}
    generic = {"burger","pizza","tacos","taco","tortas","torta","wings",
               "alitas","sushi","coffee","cafe","grill","cocina","comida"}

    # Sub-entidades de comida rápida: son palabras de marca, NO ignorar
    FAST_FOOD_SUBS = {
        "postres","desayunos","pollos","pollo","chicken",
        "vegetal","mccafe","cafe","express","drive","thru","junior"
    }

    # Cadenas de comida rápida conocidas: su nombre es 1 sola palabra clave
    FAST_FOOD_CHAINS = {
        "mcdonald","mcdonalds","burgerkng","burgerking","kfc","subway",
        "dominos","pizzahut","wendys","popeyes","chickfila","tacobell"
    }

    words_raw = norm(name).split()
    # También preservar palabras originales con tildes para word_ilike
    raw_words_orig = re.split(r'\s+', name.strip().lower())

    # Detectar si el nombre contiene una cadena + sub-entidad
    chain_word = next((w for w in words_raw if any(c in w for c in FAST_FOOD_CHAINS)), None)
    sub_word   = next((w for w in words_raw if w in FAST_FOOD_SUBS), None)

    # Si el chain_word coincide con alguna variante más corta en FAST_FOOD_CHAINS,
    # usar la más corta para capturar variantes con apóstrofes o espacios.
    # Ej: "mcdonalds" → "mcdonald" (matchea McDonald's, McDonald, Mc Donalds, Mcdonalds)
    # Esto evita que ilike '%mcdonalds%' falle con "McDonald's" (donde el apóstrofe separa 'd' y 's')
    if chain_word:
        # Buscar la variante más corta del set que esté contenida en chain_word
        shorter_variants = [c for c in FAST_FOOD_CHAINS if c in chain_word and len(c) < len(chain_word)]
        if shorter_variants:
            chain_word = min(shorter_variants, key=len)

    if chain_word and sub_word:
        return [chain_word, sub_word]

    # Flujo normal: palabras sin stopwords geo, longitud > 1
    # Usar palabras originales (con tildes) cuando estén disponibles
    words = []
    for i, w in enumerate(words_raw):
        if len(w) > 1 and w not in GEO_STOP:
            orig = re.sub(r'[^\w]', '', raw_words_orig[i]) if i < len(raw_words_orig) else w
            words.append(orig if orig else w)

    # Eliminar artículos iniciales
    while words and norm(words[0]) in func:
        words = words[1:]

    result = words[:3]

    # Si solo queda 1 palabra genérica, incluir también palabra anterior
    if len(result) == 1 and norm(result[0]) in generic:
        idx = next((i for i,w in enumerate(words_raw) if w == norm(result[0])), -1)
        if idx > 0:
            prev = []
            for j in range(max(0,idx-2), idx):
                if words_raw[j] not in GEO_STOP:
                    orig = re.sub(r'[^\w]', '', raw_words_orig[j]) if j < len(raw_words_orig) else words_raw[j]
                    prev.append(orig if orig else words_raw[j])
            result = prev + result

    return result[:3]

def build_subgroups(bad_members):
    if not bad_members: return []

    FAST_FOOD_SUBS = {
        "postres","desayunos","pollos","pollo","chicken",
        "vegetal","mccafe","cafe","express","drive","thru","junior"
    }

    def has_sub_entity(name):
        return any(w in FAST_FOOD_SUBS for w in norm(name).split())

    used = [False]*len(bad_members)
    subgroups = []
    for i, a in enumerate(bad_members):
        if used[i]: continue
        wa = set(brand_words(a.get("app_name","")))
        ak = set(extract_addr_keys(a.get("app_address","")))
        grp = [a]; used[i] = True
        # Marca genérica: 1 sola palabra → separar por dirección siempre
        generic_brand = len(wa) <= 1
        for j, b in enumerate(bad_members):
            if used[j] or i==j: continue
            wb = set(brand_words(b.get("app_name","")))
            if not wa or not wb: continue
            union = wa|wb; inter = wa&wb
            if not union or len(inter)/len(union) < 0.60: continue
            # Separar por dirección si: marca genérica O tiene sub-entidad
            if generic_brand or has_sub_entity(a.get("app_name","")) or has_sub_entity(b.get("app_name","")):
                bk = set(extract_addr_keys(b.get("app_address","")))
                if ak and bk and not ak & bk:
                    continue  # direcciones distintas → subgrupos separados
            grp.append(b); used[j] = True
        grp.sort(key=lambda m: m.get("item_index",""))
        subgroups.append({
            "rep_name":   grp[0].get("app_name",""),
            "rep_addr":   grp[0].get("app_address",""),
            "members":    [m.get("item_index","") for m in grp],
            "correction": "", "is_new": False,
        })
    return subgroups

def extract_addr_keys(addr):
    """
    Extrae claves de dirección: número de calle + primera palabra larga no geográfica.
    Devuelve palabras en su forma ORIGINAL (con tildes/Ñ) para que addr_ilike_safe
    pueda detectar caracteres problemáticos y cortar antes de ellos.
    Ej: 'Cristóbal Colón 4455' → ['4455', 'Cristóbal']
    """
    addr_geo_stop = {
        "calle","avenida","av","blvd","boulevard","carretera","carr",
        "sur","norte","oriente","poniente","ote","pte","col","colonia",
        "fracc","fraccionamiento","local","plaza","paseo","zona","barrio",
        "mexico","latam","cp","sin","nombre","entre",
    }
    # Dividir por comas y espacios preservando el original
    raw_words = re.split(r"[,\s]+", addr.strip())
    keys = []
    for w in raw_words:
        wc = re.sub(r"[^\w]", "", w)
        wc_norm = norm(wc)
        if wc_norm.isdigit() and 2 <= len(wc_norm) <= 5:
            keys.append(wc_norm)  # números siempre normalizados
        elif len(wc_norm) > 4 and wc_norm.isalpha() and wc_norm not in addr_geo_stop:
            keys.append(w.strip('.,;'))  # palabra original con tildes
        if len(keys) >= 2:
            break
    return keys

ACCENT_MAP = [
    ("a", "á"), ("e", "é"), ("i", "í"), ("o", "ó"), ("u", "ú"), ("n", "ñ"),
]

def accent_variants(word):
    """
    Genera variantes de una palabra con y sin tildes.
    'burreria' → ['burreria', 'burrería']
    'burrerıa'  → ['burrería', 'burreria']
    """
    variants = {word}
    for plain, accented in ACCENT_MAP:
        if plain in word:
            variants.add(word.replace(plain, accented, 1))
        if accented in word:
            variants.add(word.replace(accented, plain, 1))
    return list(variants)

def word_ilike(field, word):
    """Genera condición ilike robusta a tildes y Ñ.
    Corta antes del primer carácter problemático (tilde, ñ, doble letra).
    Si la palabra ya viene normalizada (sin tildes), igual aplica el corte por dobles.
    """
    # Usar addr_ilike_safe que ya implementa el corte correcto
    return addr_ilike_safe(field, word)

def addr_ilike_safe(field, word):
    """
    Busca una palabra en dirección de forma robusta a tildes, Ñ y letras dobles.
    Estrategia simple: trabajar sobre el original (con tildes/Ñ) y cortar antes
    del primer carácter problemático.
    Problemáticos: á é í ó ú ñ y letras dobles (ll, rr, cc).
    Ej: 'Colón'   → '%col%'   (corta antes de ó)
        'Castaño' → '%casta%' (corta antes de ñ)
        'Mackenna'→ '%mack%'  (corta antes de doble nn? → usa 'mack')
        'Irarrázaval' → '%irar%' (corta antes de rr)
    Si el prefijo es muy corto (<3), usa el sufijo después del carácter problemático.
    """
    # Trabajar sobre el original para detectar caracteres con tilde/Ñ
    original = word.strip().lower()
    w_norm = norm(word)
    if not w_norm: return f"{field} ilike '%{word}%'"
    if w_norm.isdigit() or len(w_norm) <= 3: return f"{field} ilike '%{w_norm}%'"

    # Caracteres problemáticos y su posición en el original
    PROBLEMATIC = set('áéíóúàèìòùâêîôûäëïöüñ')
    DOUBLES = ['rr', 'll', 'cc', 'nn', 'ss', 'pp']

    # Buscar posición del primer carácter problemático
    cut = len(original)  # por defecto no hay corte

    for i, c in enumerate(original):
        if c in PROBLEMATIC:
            cut = i
            break

    # Buscar dobles letras y tomar la posición de la primera
    for double in DOUBLES:
        idx = original.find(double)
        if idx >= 0 and idx < cut:
            cut = idx

    prefix = norm(original[:cut])   # normalizar el prefijo (quita tildes si hubiera)
    suffix = norm(original[cut+1:]) if cut < len(original) else ''

    if len(prefix) >= 3:
        return f"{field} ilike '%{prefix}%'"
    if len(suffix) >= 3:
        return f"{field} ilike '%{suffix}%'"

    # Sin corte útil: generar variantes con/sin tilde para cubrir ambos casos en la BD
    variants = accent_variants(w_norm)
    if len(variants) > 1:
        conditions = " or ".join(f"{field} ilike '%{v}%'" for v in sorted(variants))
        return f"({conditions})"
    return f"{field} ilike '%{w_norm}%'"

def get_pg_table():
    """Retorna la tabla correcta según el tipo de revisión de la sesión."""
    rt = session.get("review_type", "stores_restaurant")
    if rt == "stores_retail":
        return "sales_opportunity.dim_maestra_retail"
    return "sales_opportunity.dim_maestra"

def generate_unified_sql(bad_members, country="mx"):
    """
    Genera SQL usando:
    - brand_words(name): palabras de marca sin sufijos geográficos
    - extract_addr_keys(addr): número de calle + palabra distintiva de dirección
    - accent_variants: maneja palabras con/sin tilde automáticamente
    Para cadenas con sub-entidad (McDonald's Postres), busca por cadena Y sub-entidad
    con OR entre ellas para mayor flexibilidad.
    """
    if not bad_members: return ""
    seen, blocks = set(), []

    FAST_FOOD_SUBS = {
        "postres","desayunos","pollos","pollo","chicken",
        "vegetal","mccafe","cafe","express","drive","thru","junior"
    }

    for m in bad_members:
        name = m.get("app_name","")
        addr = m.get("app_address","")

        bwords = brand_words(name)
        if len(bwords) < 1:
            bwords = [w for w in norm(name).split() if len(w) > 3][:2]

        sql_bwords = [w for w in bwords if len(w) > 2] or bwords

        # Detectar si hay sub-entidad entre las brand_words
        sub = next((w for w in sql_bwords if w in FAST_FOOD_SUBS), None)
        chain_words = [w for w in sql_bwords if w not in FAST_FOOD_SUBS]

        if sub and chain_words:
            akeys = extract_addr_keys(addr)
            key = "_".join(chain_words) + "_" + sub + "_" + "_".join(akeys)
            if key in seen: continue
            seen.add(key)
            # Usar nombre original para detectar Ñ correctamente
            chain_cond = " and ".join(word_ilike("app_name", w) for w in chain_words)
            sub_cond   = word_ilike("app_name", sub)
            name_cond  = f"({chain_cond}\n     and {sub_cond})"
        else:
            key = "_".join(bwords)
            if key in seen or not bwords: continue
            seen.add(key)
            # Pasar nombre original para detectar Ñ; si hay una sola bword usar el nombre completo
            if len(sql_bwords) == 1:
                name_cond = word_ilike("app_name", sql_bwords[0])
            else:
                name_cond = " and ".join(word_ilike("app_name", w) for w in sql_bwords)

        # Condición de dirección — OR entre palabras clave (más permisivo, evita falsos negativos)
        ADDR_BLACKLIST = {
            "no", "sin", "especificado", "especificada", "null", "none",
            "nulo", "nula", "vacio", "vacia", "nd", "na", "nr", "señala",
            "indica", "definido", "definida", "disponible", "dato", "datos"
        }
        akeys = extract_addr_keys(addr)
        # Filtrar claves que sean palabras inútiles
        akeys = [k for k in akeys if norm(k) not in ADDR_BLACKLIST and len(norm(k)) >= 3]
        if akeys:
            addr_parts = [addr_ilike_safe("app_address", w) for w in akeys]
            addr_cond = " or ".join(addr_parts) if len(addr_parts) > 1 else addr_parts[0]
            blocks.append(f"    ({name_cond}\n     and ({addr_cond}))")
        else:
            # Sin dirección útil: solo condición de nombre
            name_key = "_".join(sql_bwords) + "_noaddr"
            if name_key not in seen:
                seen.add(name_key)
                blocks.append(f"    ({name_cond})")

    if not blocks: return ""

    # Limitar a 8 bloques máximo para evitar queries demasiado largas
    if len(blocks) > 8:
        blocks = blocks[:8]

    comment = "; ".join(m.get("app_name","")[:35] for m in bad_members[:5])
    if len(bad_members) > 5: comment += "..."

    return (
        f"-- {len(bad_members)} store(s): {comment}\n"
        f"select\n  cluster_index, store_id, cluster_name, main_chain,\n"
        f"  app_name, app_address, app_longitude, app_latitude\n"
        f"from {get_pg_table()}\n"
        f"where country = '{country}'\n"
        f"  and (\n" + "\n    or\n".join(blocks) + "\n  )\n"
        f"order by cluster_index;"
    )


# ── build clusters ────────────────────────────────────────────────────────────

def build_clusters(df, filename=""):
    grouped = {}
    for _, row in df.iterrows():
        r = {k:("" if pd.isna(v) else str(v)) for k,v in row.items()}
        cf = effective_cluster(r)
        if cf: grouped.setdefault(cf,[]).append(r)

    clusters = []; threshold = get_threshold()
    fb = load_feedback()

    for cf, members in grouped.items():
        anchor = next((m for m in members if m.get("item_index","").strip()==cf), None)
        if anchor is None:
            anchor = sorted(members, key=lambda m: m.get("item_index",""))[0]
        rest = sorted([m for m in members if m is not anchor],
                      key=lambda m: norm(m.get("app_name","")))
        anchor_name = anchor.get("app_name","")
        anchor_addr = anchor.get("app_address","")
        scores = batch_semantic_sim(anchor_name, anchor_addr, rest)

        members_out = [{
            "item_index": anchor.get("item_index","").strip(),
            "new_cluster": anchor.get("new_cluster",cf),
            "cluster_index": anchor.get("cluster_index",""),
            "app_name": anchor_name, "app_address": anchor_addr,
            "store_id": anchor.get("store_id",""),
            "scraper_source": anchor.get("scraper_source",""),
            "is_anchor": True, "revision": 1, "score": None,
            "mem_hit": None,
            **{k: anchor.get(k,"") for k in ["cluster_name","main_chain","country",
               "cluster_latitude","cluster_longitude","app_latitude","app_longitude",
               "cluster_ciudad","cluster_estado","app_phone_1","app_phone_2","app_phone_3"]}
        }]

        for i, m in enumerate(rest):
            ii    = m.get("item_index","").strip()
            score = scores[i]
            rev   = 1 if score >= threshold else 0

            # Consultar memoria interna
            mem_hit = mem_lookup_item(ii)
            if not mem_hit:
                mem_hit, _ = mem_lookup_similar(m.get("app_name",""), anchor_name)

            members_out.append({
                "item_index": ii,
                "new_cluster": m.get("new_cluster",cf),
                "cluster_index": m.get("cluster_index",""),
                "app_name": m.get("app_name",""),
                "app_address": m.get("app_address",""),
                "store_id": m.get("store_id",""),
                "scraper_source": m.get("scraper_source",""),
                "is_anchor": False, "revision": rev, "score": score,
                "mem_hit": mem_hit,
                **{k: m.get(k,"") for k in ["cluster_name","main_chain","country",
                   "cluster_latitude","cluster_longitude","app_latitude","app_longitude",
                   "cluster_ciudad","cluster_estado","app_phone_1","app_phone_2","app_phone_3"]}
            })

        clusters.append({
            "cluster_id": cf, "anchor_name": anchor_name, "anchor_addr": anchor_addr,
            "count": len(members_out),
            "ok_count": sum(1 for m in members_out if m["revision"]==1),
            "bad_count": sum(1 for m in members_out if m["revision"]==0),
            "members": members_out, "sql": "", "bd_results": [],
            "corrections": {}, "threshold_used": threshold,
        })

    clusters.sort(key=lambda c: c["cluster_id"])
    return clusters, {"threshold": threshold, "fb_stats": fb["stats"],
                      "model_ready": _model_ready, "pg_ready": pg_is_configured()}


def ext_corr_add(record):
    """Agrega una corrección externa (store no en el archivo)."""
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO external_corrections
          (store_id, item_index, cluster_index, cluster_name, cluster_address,
           app_name, app_address, scraper_source, correction, added_at, file_name, is_new)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        record.get("store_id",""),   record.get("item_index",""),
        record.get("cluster_index",""), record.get("cluster_name",""),
        record.get("cluster_address",""), record.get("app_name",""),
        record.get("app_address",""), record.get("scraper_source",""),
        record.get("correction",""),
        time.time(), record.get("file_name",""),
        1 if record.get("is_new") else 0
    ))
    conn.commit(); conn.close()

def ext_corr_list():
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute("""
        SELECT id, store_id, item_index, cluster_index, cluster_name,
               cluster_address, app_name, app_address, scraper_source,
               correction, added_at, file_name
        FROM external_corrections ORDER BY added_at DESC
    """).fetchall()
    conn.close()
    cols = ["id","store_id","item_index","cluster_index","cluster_name",
            "cluster_address","app_name","app_address","scraper_source","correction","added_at","file_name"]
    return [dict(zip(cols,r)) for r in rows]

def ext_corr_delete(rec_id):
    conn = sqlite3.connect(DB_FILE)
    conn.execute("DELETE FROM external_corrections WHERE id=?", (rec_id,))
    conn.commit(); conn.close()

def save_reviewed_clusters_dishes(groups, filename):
    """Guarda grupos de dishes en reviewed_clusters."""
    conn = sqlite3.connect(DB_FILE)
    for g in groups:
        conn.execute("""
            INSERT INTO reviewed_clusters
              (cluster_index, cluster_name, had_errors, corrections_count, reviewed_at, file_name)
            VALUES (?,?,?,?,?,?)
        """, (g["group_id"], "", 1 if g["revision"] in (0,2) else 0,
              0, time.time(), filename))
    conn.commit(); conn.close()

def save_reviewed_file(file_name, review_type, total, ok, bad, incomplete=0):
    """Guarda un resumen del archivo revisado al descargar."""
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO reviewed_files (file_name, review_type, total, ok, bad, incomplete, reviewed_at)
        VALUES (?,?,?,?,?,?,?)
    """, (file_name, review_type, total, ok, bad, incomplete, time.time()))
    conn.commit(); conn.close()

def get_reviewed_files(review_type=None):
    """Retorna historial de archivos revisados."""
    conn = sqlite3.connect(DB_FILE)
    if review_type:
        rows = conn.execute("""
            SELECT file_name, review_type, total, ok, bad, incomplete, reviewed_at
            FROM reviewed_files WHERE review_type=?
            ORDER BY reviewed_at DESC
        """, (review_type,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT file_name, review_type, total, ok, bad, incomplete, reviewed_at
            FROM reviewed_files ORDER BY reviewed_at DESC
        """).fetchall()
    conn.close()
    cols = ["file_name","review_type","total","ok","bad","incomplete","reviewed_at"]
    return [dict(zip(cols,r)) for r in rows]

def save_reviewed_clusters(clusters, filename, review_type=None):
    """Guarda qué clusters pasaron por revisión al exportar."""
    rt = review_type or session.get("review_type","stores_restaurant")
    conn = sqlite3.connect(DB_FILE)
    for cl in clusters:
        had_errors = cl.get("bad_count",0) > 0
        corr_count = len(cl.get("corrections",{}))
        conn.execute("""
            INSERT INTO reviewed_clusters
              (cluster_index, cluster_name, had_errors, corrections_count, reviewed_at, file_name, review_type)
            VALUES (?,?,?,?,?,?,?)
        """, (
            cl["cluster_id"],
            cl.get("anchor_name",""),
            1 if had_errors else 0,
            corr_count,
            time.time(),
            filename,
            rt
        ))
    conn.commit(); conn.close()

def progress_stats(table="sales_opportunity.dim_maestra", country="mx", review_type="stores_restaurant"):
    """Cruza la memoria con la BD para calcular progreso.
    Filtra por review_type (stores_restaurant vs stores_retail) vía join con reviewed_files."""
    if not pg_is_configured():
        return None, "BD no configurada"

    # Filtro de country
    country_filter = f"WHERE country = '{country}'" if country != "all" else ""

    # Clusters y tiendas totales en BD
    sql_total = f"""
        SELECT
            COUNT(DISTINCT cluster_index) as total_clusters,
            COUNT(*) as total_stores
        FROM {table}
        {country_filter}
    """
    rows, err = pg_query(sql_total, timeout_ms=30000)
    if err:
        if "does not exist" in str(err) or "doesn't exist" in str(err):
            return None, f"La tabla '{table}' no existe en la BD"
        return None, err
    total_clusters = rows[0]["total_clusters"] if rows else 0
    total_stores   = rows[0]["total_stores"]   if rows else 0

    conn = sqlite3.connect(DB_FILE)
    # Clusters revisados y corregidos filtrados por review_type
    reviewed_clusters = conn.execute("""
        SELECT COUNT(DISTINCT cluster_index) FROM reviewed_clusters
        WHERE review_type = ?
    """, (review_type,)).fetchone()[0]
    corrected_clusters = conn.execute("""
        SELECT COUNT(DISTINCT cluster_index) FROM reviewed_clusters
        WHERE corrections_count > 0 AND review_type = ?
    """, (review_type,)).fetchone()[0]

    # Correcciones propias del archivo (del subgrupo) — join con reviewed_files para filtrar por review_type
    # Nota: corrections.file_name debe existir en reviewed_files con el review_type correcto
    own_migrated = conn.execute("""
        SELECT COUNT(*) FROM corrections c
        WHERE c.is_new = 0
        AND EXISTS (
            SELECT 1 FROM reviewed_files rf
            WHERE rf.file_name = c.file_name AND rf.review_type = ?
        )
    """, (review_type,)).fetchone()[0]
    own_new = conn.execute("""
        SELECT COUNT(*) FROM corrections c
        WHERE c.is_new = 1
        AND EXISTS (
            SELECT 1 FROM reviewed_files rf
            WHERE rf.file_name = c.file_name AND rf.review_type = ?
        )
    """, (review_type,)).fetchone()[0]

    # Correcciones externas (stores que no estaban en el archivo, arrastradas desde BD)
    ext_migrated = conn.execute("""
        SELECT COUNT(*) FROM external_corrections ec
        WHERE COALESCE(ec.is_new, 0) = 0
        AND EXISTS (
            SELECT 1 FROM reviewed_files rf
            WHERE rf.file_name = ec.file_name AND rf.review_type = ?
        )
    """, (review_type,)).fetchone()[0]
    ext_new = conn.execute("""
        SELECT COUNT(*) FROM external_corrections ec
        WHERE COALESCE(ec.is_new, 0) = 1
        AND EXISTS (
            SELECT 1 FROM reviewed_files rf
            WHERE rf.file_name = ec.file_name AND rf.review_type = ?
        )
    """, (review_type,)).fetchone()[0]

    recent = conn.execute("""
        SELECT cluster_index, cluster_name, had_errors, corrections_count, reviewed_at, file_name
        FROM reviewed_clusters WHERE review_type = ?
        ORDER BY reviewed_at DESC LIMIT 20
    """, (review_type,)).fetchall()
    conn.close()

    # Totales compuestos (para compatibilidad con UI anterior)
    migrated = own_migrated + ext_migrated
    new_cluster = own_new + ext_new

    pct = round(reviewed_clusters / int(total_clusters) * 100, 1) if total_clusters else 0
    return {
        "table":              table,
        "country":            country,
        "review_type":        review_type,
        "total_clusters":     int(total_clusters),
        "total_stores":       int(total_stores),
        "reviewed":           reviewed_clusters,
        "corrected":          corrected_clusters,
        # Compatibilidad con UI anterior
        "migrated_stores":    migrated,
        "new_cluster_stores": new_cluster,
        # Desglose nuevo (propias vs externas, migradas vs nuevas)
        "own_migrated":       own_migrated,
        "own_new":            own_new,
        "ext_migrated":       ext_migrated,
        "ext_new":            ext_new,
        "own_total":          own_migrated + own_new,
        "ext_total":          ext_migrated + ext_new,
        "pending":            max(0, int(total_clusters) - reviewed_clusters),
        "pct":                pct,
        "recent": [
            {"cluster_index":r[0],"cluster_name":r[1],"had_errors":bool(r[2]),
             "corrections_count":r[3],"reviewed_at":r[4],"file_name":r[5]}
            for r in recent
        ]
    }, None

# ── dishes (revisión de platos) ───────────────────────────────────────────────

def pairwise_sim_semantic(names, descs):
    """Similitud semántica promedio entre todos los pares. Sin fallback Jaccard."""
    texts = [f"{n} {d}".strip() for n, d in zip(names, descs)]
    if len(texts) < 2:
        return None  # grupo de 1 → incompleto
    m = get_model()
    if not m:
        return None  # sin modelo → no evaluar
    from sentence_transformers import util
    embs = m.encode(texts, convert_to_tensor=True, show_progress_bar=False)
    sims = []
    for i in range(len(texts)):
        for j in range(i+1, len(texts)):
            sims.append(float(util.cos_sim(embs[i], embs[j])))
    return round(sum(sims)/len(sims), 3)

def detect_error_categories(members):
    """
    Detecta categorías de error automáticamente.
    Retorna lista de categorías: price, name, description, no_data
    """
    categories = []

    # Error de precio: diferencia > 20% entre min y max
    prices = []
    for m in members:
        try:
            p = float(str(m.get("price","")).replace(",",".").strip())
            if p > 0: prices.append(p)
        except: pass
    if len(prices) >= 2:
        mn, mx = min(prices), max(prices)
        if mx > 0 and (mx - mn) / mx > 0.20:
            categories.append("price")

    # Error de nombre: palabra clave distintiva en uno que no está en otros
    names = [norm(m.get("product_name","")) for m in members]
    all_words = [set(w for w in n.split() if len(w) > 3) for n in names]
    for i, wi in enumerate(all_words):
        for j, wj in enumerate(all_words):
            if i == j: continue
            exclusive = wi - wj  # palabras de i que no están en j
            if exclusive:
                categories.append("name")
                break
        if "name" in categories: break

    # Error de descripción: baja similitud entre descripciones (si existen)
    descs = [m.get("product_description","").strip() for m in members]
    has_descs = [d for d in descs if d]
    if len(has_descs) >= 2:
        m_model = get_model()
        if m_model:
            from sentence_transformers import util
            embs = m_model.encode(has_descs, convert_to_tensor=True, show_progress_bar=False)
            sims = []
            for i in range(len(has_descs)):
                for j in range(i+1, len(has_descs)):
                    sims.append(float(util.cos_sim(embs[i], embs[j])))
            avg = sum(sims)/len(sims) if sims else 1.0
            if avg < 0.70:
                categories.append("description")
        elif not has_descs:
            categories.append("no_data")
    elif not has_descs:
        categories.append("no_data")

    return list(set(categories))

def build_dish_groups(df, filename="", threshold=0.75):
    """Construye grupos de platos a partir del archivo de revisión."""
    df = df.fillna("")
    groups = {}
    for _, row in df.iterrows():
        r = {k: str(v) for k, v in row.items()}
        gid = r.get("grupo_pn9_max", "").strip()
        if not gid: continue
        groups.setdefault(gid, []).append(r)

    result = []
    for gid, members in groups.items():
        names = [m.get("product_name","") for m in members]
        descs = [m.get("product_description","") for m in members]
        score = pairwise_sim_semantic(names, descs)

        # revision: 2=incompleto (1 solo miembro o sin modelo), 1=correcto, 0=incorrecto
        if score is None:
            revision = 2  # incompleto — grupo de 1 o sin modelo
        elif score >= threshold:
            revision = 1
        else:
            revision = 0

        # Categorías de error (solo para incorrectos)
        error_cats = detect_error_categories(members) if revision in (0, 2) else []

        result.append({
            "group_id":    gid,
            "score":       score,
            "revision":    revision,
            "count":       len(members),
            "error_cats":  error_cats,
            "fusion_with": "",  # grupo_pn9_max con el que fusionar (si revision==2)
            "members": [{
                "product_id":          m.get("product_id",""),
                "scraper_source":      m.get("scraper_source",""),
                "product_name":        m.get("product_name",""),
                "product_description": m.get("product_description",""),
                "price":               m.get("price",""),
                "cluster_index":       m.get("cluster_index",""),
            } for m in members]
        })

    result.sort(key=lambda g: g["group_id"])
    ok   = sum(1 for g in result if g["revision"]==1)
    bad  = sum(1 for g in result if g["revision"]==0)
    inc  = sum(1 for g in result if g["revision"]==2)
    return result, {
        "total_groups": len(result), "ok": ok, "bad": bad, "incomplete": inc,
        "threshold": threshold, "model_ready": _model_ready
    }



def save_store_pair(app_name_a, app_address_a, cluster_index_a,
                    app_name_b, app_address_b, cluster_index_b,
                    label=0, source="correction", file_name=""):
    """Guarda un par etiquetado de stores para futuro fine-tuning del algoritmo de clusterización.
    label=1: mismo local físico, label=0: locales distintos.
    """
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO store_pairs
          (app_name_a, app_address_a, cluster_index_a,
           app_name_b, app_address_b, cluster_index_b,
           label, source, added_at, file_name)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (app_name_a, app_address_a, cluster_index_a,
          app_name_b, app_address_b, cluster_index_b,
          label, source, time.time(), file_name))
    conn.commit(); conn.close()

def save_dish_pair(name_a, desc_a, name_b, desc_b, label, source="revision", file_name=""):
    """Guarda un par etiquetado de dishes para futuro fine-tuning."""
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO dish_pairs (name_a, desc_a, name_b, desc_b, label, source, added_at, file_name)
        VALUES (?,?,?,?,?,?,?,?)
    """, (name_a, desc_a, name_b, desc_b, label, source, time.time(), file_name))
    conn.commit(); conn.close()

def compute_group_similarity(g_target, g_other):
    """
    Calcula similitud entre dos grupos de dishes.
    Combina nombre (70%) y descripción (30%).
    """
    # Texto representativo de cada grupo (primer miembro)
    name_a = g_target["members"][0].get("product_name","") if g_target["members"] else ""
    desc_a = g_target["members"][0].get("product_description","") if g_target["members"] else ""
    name_b = g_other["members"][0].get("product_name","") if g_other["members"] else ""
    desc_b = g_other["members"][0].get("product_description","") if g_other["members"] else ""

    m = get_model()
    if m:
        from sentence_transformers import util
        texts = [name_a, name_b, desc_a or name_a, desc_b or name_b]
        embs  = m.encode(texts, convert_to_tensor=True, show_progress_bar=False)
        sim_name = float(util.cos_sim(embs[0], embs[1]))
        sim_desc = float(util.cos_sim(embs[2], embs[3]))
    else:
        sim_name = jaccard_sim(name_a, name_b)
        sim_desc = jaccard_sim(desc_a or name_a, desc_b or name_b)

    return round(sim_name * 0.7 + sim_desc * 0.3, 3)

# ── rutas ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/session/restore")
def session_restore():
    """Devuelve el estado de la sesión activa si existe."""
    if session.get("clusters"):
        clusters = session.get("clusters",[])
        return jsonify({
            "type":     "stores",
            "filename": session.get("filename",""),
            "clusters": clusters,
            "country":  session.get("country","mx"),
            "stats": {
                "total_rows":    sum(len(c["members"]) for c in clusters),
                "total_clusters":len(clusters),
                "total_ok":      sum(c["ok_count"] for c in clusters),
                "total_bad":     sum(c["bad_count"] for c in clusters),
            }
        })
    elif session.get("dishes_groups"):
        groups = session.get("dishes_groups",[])
        ok  = sum(1 for g in groups if g["revision"]==1)
        bad = sum(1 for g in groups if g["revision"]==0)
        inc = sum(1 for g in groups if g["revision"]==2)
        return jsonify({
            "type":     "dishes",
            "filename": session.get("dishes_filename",""),
            "groups":   groups,
            "meta":     {"total_groups":len(groups),"ok":ok,"bad":bad,
                         "incomplete":inc,"threshold":0.75,"model_ready":_model_ready}
        })
    return jsonify({"type": None})

@app.route("/model_status")
def model_status():
    get_model()
    return jsonify({"ready":_model_ready,"error":_model_error,"model":MODEL_NAME})

@app.route("/pg_status")
def pg_status():
    if not pg_is_configured():
        return jsonify({"ok":False,"error":"No configurado (.env)"})
    try:
        import psycopg2
        cfg = get_pg_config()
        conn = psycopg2.connect(**cfg, connect_timeout=3)  # 3 segundos máximo
        conn.close()
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@app.route("/cluster_search", methods=["POST"])
def cluster_search():
    """Busca clusters en dim_maestra con filtros separados.
    - name_query: palabras buscadas en app_name (AND entre palabras)
    - addr_query: palabras buscadas en app_address (AND entre palabras)
    - size_filter: 'all' | 'multi' (>=2 stores) | 'small' (1-2 stores)
    - cluster_index: si viene, se busca exacto por cluster_index (ignora otros filtros)
    Devuelve clusters únicos con sus miembros para trabajar sobre ellos
    independientemente del archivo en revisión.
    Compatibilidad: si llega 'query' (versión vieja), se usa como name_query.
    """
    if not pg_is_configured():
        return jsonify({"error": "BD no configurada"}), 400
    d = request.json or {}
    country = (d.get("country") or session.get("country","mx")).strip()
    name_q = (d.get("name_query") or d.get("query") or "").strip()
    addr_q = (d.get("addr_query") or "").strip()
    cluster_idx = (d.get("cluster_index") or "").strip()
    size_filter = (d.get("size_filter") or "all").strip()
    limit = max(10, min(200, int(d.get("limit") or 100)))

    table = get_pg_table()

    def _words(s):
        s_norm = re.sub(r'[^\w\s]', ' ', s.lower()).strip()
        return [w for w in s_norm.split() if len(w) >= 3]

    # Modo cluster_index exacto: bypass otros filtros
    if cluster_idx:
        # Sanitizar (solo caracteres seguros)
        ci_safe = re.sub(r"[^\w:.\-]", "", cluster_idx)
        if not ci_safe:
            return jsonify({"error": "cluster_index inválido"}), 400
        sql = f"""
            SELECT DISTINCT cluster_index, cluster_name, cluster_address, country
            FROM {table}
            WHERE country = '{country}' AND cluster_index = '{ci_safe}'
            LIMIT 1
        """
        rows, err = pg_query(sql)
        if err: return jsonify({"error": err}), 400
        if not rows: return jsonify({"clusters": [], "count": 0, "total": 0, "limited": False})
    else:
        name_words = _words(name_q)
        addr_words = _words(addr_q)
        if not name_words and not addr_words:
            return jsonify({"error": "Especifica al menos un término en nombre o dirección (mín 3 chars)"}), 400

        conditions = []
        for w in name_words[:4]:
            conditions.append(f"app_name ilike '%{w}%'")
        for w in addr_words[:4]:
            conditions.append(f"app_address ilike '%{w}%'")
        where = " AND ".join(conditions)

        # Filtro de tamaño: aplicado en HAVING tras agrupar
        size_having = ""
        if size_filter == "multi":
            size_having = "HAVING COUNT(*) >= 2"
        elif size_filter == "small":
            size_having = "HAVING COUNT(*) <= 2"

        # Primero contamos total de clusters que cumplen filtros (para mostrar al usuario si pegó al límite)
        count_sql = f"""
            SELECT COUNT(*) AS total FROM (
                SELECT cluster_index
                FROM {table}
                WHERE country = '{country}' AND ({where})
                GROUP BY cluster_index
                {size_having}
            ) t
        """
        count_rows, count_err = pg_query(count_sql, timeout_ms=15000)
        total_count = int(count_rows[0]["total"]) if count_rows and not count_err else None

        # Query principal: agrupada por cluster_index, con tamaño calculado
        sql = f"""
            SELECT cluster_index,
                   MAX(cluster_name)    AS cluster_name,
                   MAX(cluster_address) AS cluster_address,
                   COUNT(*)             AS n_stores
            FROM {table}
            WHERE country = '{country}' AND ({where})
            GROUP BY cluster_index
            {size_having}
            ORDER BY cluster_name
            LIMIT {limit}
        """
        rows, err = pg_query(sql)
        if err: return jsonify({"error": err}), 400
        if not rows:
            return jsonify({"clusters": [], "count": 0, "total": total_count or 0, "limited": False})

    # Traer todos los miembros de los clusters encontrados
    ci_list = ", ".join(f"'{str(r.get('cluster_index',''))}'" for r in rows)
    members_sql = f"""
        SELECT cluster_index, store_id, item_index, app_name, app_address,
               app_latitude, app_longitude, scraper_source, main_chain
        FROM {table}
        WHERE country = '{country}'
          AND cluster_index IN ({ci_list})
        ORDER BY cluster_index, store_id
    """
    members_rows, err2 = pg_query(members_sql)
    if err2:
        members_rows = []

    # Agrupar miembros por cluster
    by_cluster = {}
    for r in members_rows:
        ci = str(r.get("cluster_index",""))
        m = {k: str(v) if v is not None else "" for k, v in r.items()}
        m["is_anchor"] = (str(r.get("item_index","")) == ci)
        by_cluster.setdefault(ci, []).append(m)

    clusters = []
    for r in rows:
        ci = str(r.get("cluster_index",""))
        clusters.append({
            "cluster_id":      ci,
            "cluster_index":   ci,
            "anchor_name":     str(r.get("cluster_name","")),
            "anchor_address":  str(r.get("cluster_address","")),
            "members":         by_cluster.get(ci, []),
            "ok_count":        0,
            "bad_count":       len(by_cluster.get(ci, [])),
        })

    # total = cuántos clusters cumplen filtros en BD (puede ser > limit)
    # limited = True si pegamos al techo del limit
    total = total_count if not cluster_idx and total_count is not None else len(clusters)
    limited = (not cluster_idx) and (total_count is not None) and (total_count > len(clusters))
    return jsonify({"clusters": clusters, "count": len(clusters), "total": total, "limited": limited})

@app.route("/pg_query", methods=["POST"])
def run_pg_query():
    """
    Ejecuta la SQL de búsqueda, obtiene los cluster_index que matchean,
    luego trae TODOS los miembros de esos clusters para poder evaluar si están limpios.
    """
    data = request.json
    sql  = data.get("sql","").strip()
    if not sql: return jsonify({"error":"SQL vacía"}), 400
    if not pg_is_configured():
        return jsonify({"error":"BD no configurada. Revisa el archivo .env"}), 400

    # 1. Query original — encontrar clusters que matchean
    rows, err = pg_query(sql)
    if err: return jsonify({"error": err}), 400

    # Obtener cluster_indexes únicos encontrados
    matched_clusters = list({str(r.get("cluster_index","")) for r in rows if r.get("cluster_index")})

    if not matched_clusters:
        return jsonify({"clusters_found": [], "row_count": 0})

    # 2. Segunda query — traer TODOS los miembros de esos clusters
    ci_list = ", ".join(f"'{ci}'" for ci in matched_clusters)
    # Detectar country de la primera query
    country = "mx"
    for line in sql.lower().split("\n"):
        if "country" in line and "=" in line:
            m = re.search(r"country\s*=\s*'(\w+)'", line)
            if m: country = m.group(1); break

    full_sql = f"""select
  cluster_index, store_id, item_index, cluster_name, main_chain,
  app_name, app_address, app_longitude, app_latitude, scraper_source
from {get_pg_table()}
where country = '{country}'
  and cluster_index in ({ci_list})
order by cluster_index, store_id"""

    all_rows, err2 = pg_query(full_sql)
    if err2:
        # Fallback: usar los resultados originales si la segunda query falla
        all_rows = rows

    # Agrupar por cluster_index mostrando todos los miembros
    by_cluster = {}
    for row in all_rows:
        ci = str(row.get("cluster_index",""))
        m = {k: str(v) if v is not None else "" for k, v in row.items()}
        m["is_anchor"] = (str(row.get("item_index","")) == ci)
        by_cluster.setdefault(ci, []).append(m)

    clusters_found = [
        {
            "cluster_index": ci,
            "members": mems,
            "is_clean": False,
            "matched": ci in set(matched_clusters),  # indicar cuáles matchearon la query
        }
        for ci, mems in by_cluster.items()
    ]

    return jsonify({"clusters_found": clusters_found, "row_count": len(all_rows)})


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f: return jsonify({"error":"Sin archivo"}), 400
    ext  = Path(f.filename).suffix.lower()
    path = UPLOAD_FOLDER / f"review{ext}"
    f.save(path)
    try: df = load_file(path)
    except Exception as e: return jsonify({"error":str(e)}), 400

    review_type = request.form.get("review_type", "stores_restaurant")
    session["review_type"] = review_type

    clusters, meta = build_clusters(df, filename=f.filename)
    country = detect_country(df)

    # Intentar restaurar estado previo para este archivo
    saved_state, saved_at = load_session_state(f.filename)
    restored = False
    subgroups_restored = {}
    if saved_state:
        clusters, subgroups_restored = apply_session_state(clusters, saved_state)
        restored = True
        meta["restored"] = True
        meta["saved_at"] = saved_at

    # Marcar clusters ya revisados previamente
    reviewed_set = set()
    try:
        conn = sqlite3.connect(DB_FILE)
        rows = conn.execute("SELECT DISTINCT cluster_index FROM reviewed_clusters").fetchall()
        reviewed_set = {r[0] for r in rows}
        conn.close()
    except: pass
    for cl in clusters:
        cl["already_reviewed"] = cl["cluster_id"] in reviewed_set

    session["review_path"] = str(path)
    session["clusters"]    = clusters
    session["country"]     = country
    session["filename"]    = f.filename

    # Si el archivo ya fue marcado como revisado antes, restaurar ese estado
    already_reviewed = False
    try:
        conn = sqlite3.connect(DB_FILE)
        row = conn.execute(
            "SELECT COUNT(*) FROM reviewed_files WHERE file_name=?", (f.filename,)
        ).fetchone()
        conn.close()
        already_reviewed = row and row[0] > 0
    except: pass
    session["is_reviewed"] = already_reviewed
    session.modified = True

    return jsonify({"clusters":clusters, "meta":meta, "country":country,
                    "filename": session.get("filename",""),
                    "restored": restored,
                    "already_reviewed": already_reviewed,
                    "review_type": review_type,
                    "subgroups_restored": subgroups_restored,
                    "stats":{"total_rows":len(df),"total_clusters":len(clusters),
                             "total_ok":sum(c["ok_count"] for c in clusters),
                             "total_bad":sum(c["bad_count"] for c in clusters)}})

@app.route("/update_revisions", methods=["POST"])
def update_revisions():
    data=request.json; cid=data["cluster_id"]; revisions=data["revisions"]
    record_fb=data.get("record_feedback",True)
    clusters=session.get("clusters",[])
    sql,bad=("",0)
    for cl in clusters:
        if cl["cluster_id"]!=cid: continue
        anchor=next((m for m in cl["members"] if m["is_anchor"]),None)
        for m in cl["members"]:
            ii=m["item_index"]
            if ii not in revisions: continue
            new_rev=revisions[ii]
            if record_fb and not m["is_anchor"] and new_rev!=m["revision"]:
                record_feedback(anchor["app_name"] if anchor else "",
                                anchor["app_address"] if anchor else "",
                                m["app_name"],m["app_address"],
                                m.get("score"),m["revision"],new_rev)
            m["revision"]=new_rev
        cl["ok_count"]=sum(1 for m in cl["members"] if m["revision"]==1)
        cl["bad_count"]=sum(1 for m in cl["members"] if m["revision"]==0)
        bads=[m for m in cl["members"] if m["revision"]==0]
        cl["sql"]=generate_unified_sql(bads)
        sql,bad=cl["sql"],cl["bad_count"]
        break
    session["clusters"]=clusters; session.modified=True
    # Guardar estado en disco en tiempo real
    save_session_state(session.get("filename",""), clusters)
    fb=load_feedback()
    return jsonify({"sql":sql,"bad_count":bad,"threshold":get_threshold(),"fb_stats":fb["stats"]})

@app.route("/get_subgroups", methods=["POST"])
def get_subgroups():
    data=request.json; cid=data["cluster_id"]
    members_data=data["members_data"]; country=data.get("country") or session.get("country","mx")
    subgroups=build_subgroups(members_data)
    for sg in subgroups:
        sg_members=[m for m in members_data if m["item_index"] in sg["members"]]
        sg["sql"]=generate_unified_sql(sg_members,country=country)
    return jsonify({"subgroups":subgroups,"cluster_id":cid})

@app.route("/save_subgroups", methods=["POST"])
def save_subgroups():
    """Guarda el estado actual de subgrupos en session_state."""
    data     = request.json
    subgroups_map = data.get("subgroups", {})  # {cluster_id: [subgrupos]}
    clusters = session.get("clusters", [])
    filename = session.get("filename", "")
    if filename and clusters:
        save_session_state(filename, clusters, subgroups=subgroups_map)
    return jsonify({"ok": True})

@app.route("/set_clean", methods=["POST"])
def set_clean():
    data=request.json; cid=data["cluster_id"]
    clusters=session.get("clusters",[])
    for cl in clusters:
        if cl["cluster_id"]!=cid: continue
        for bdc in cl.get("bd_results",[]):
            if bdc["cluster_index"]==data["bd_cluster_index"]: bdc["is_clean"]=data["is_clean"]
        break
    session["clusters"]=clusters; session.modified=True
    return jsonify({"ok":True})

@app.route("/set_correction", methods=["POST"])
def set_correction():
    data=request.json; cid=data["cluster_id"]
    items=data["item_indexes"]; corr=data["correction"]
    clusters=session.get("clusters",[])
    filename=session.get("filename","")
    for cl in clusters:
        if cl["cluster_id"]!=cid: continue
        anchor=next((m for m in cl["members"] if m["is_anchor"]),None)
        for ii in items:
            if corr:
                cl["corrections"][ii]=corr
                # Guardar par etiquetado: store incorrecto vs anchor del cluster
                member=next((m for m in cl["members"] if m["item_index"]==ii),None)
                if member and anchor and corr:
                    # Par label=0: este store NO es el mismo local que el anchor
                    save_store_pair(
                        member.get("app_name",""), member.get("app_address",""), cid,
                        anchor.get("app_name",""),  anchor.get("app_address",""),  cid,
                        label=0, source="correction", file_name=filename
                    )
            else:
                cl["corrections"].pop(ii,None)
        break
    session["clusters"]=clusters; session.modified=True
    save_session_state(session.get("filename",""), clusters)
    return jsonify({"ok":True})

@app.route("/memory/search")
def memory_search():
    q=request.args.get("q","").strip()
    if not q: return jsonify({"results":[]})
    return jsonify({"results": mem_search(q)})

@app.route("/memory/stats")
def memory_stats_route():
    s=mem_stats()
    return jsonify({"total":s["total"],"unique_items":s["unique_items"],
                    "recent":[{"item_index":r[0],"app_name":r[1],
                               "correction":r[2],"reviewed_at":r[3]} for r in s["recent"]]})

@app.route("/feedback_stats")
def feedback_stats():
    fb=load_feedback()
    return jsonify({"stats":fb["stats"],"threshold":get_threshold(),
                    "threshold_adj":fb.get("threshold_adj",0.0),"pairs_count":len(fb["pairs"])})

@app.route("/reset_feedback", methods=["POST"])
def reset_feedback():
    if FEEDBACK_FILE.exists(): FEEDBACK_FILE.unlink()
    return jsonify({"ok":True})

@app.route("/export")
def export():
    review_path=session.get("review_path")
    clusters=session.get("clusters",[])
    filename=session.get("filename","archivo")
    if not review_path: return jsonify({"error":"Sin sesión"}),400
    df=load_file(review_path)
    rev_map,corr_map={},{}
    mem_records=[]
    for cl in clusters:
        anchor=next((m for m in cl["members"] if m["is_anchor"]),None)
        for m in cl["members"]:
            ii=m["item_index"]
            rev_map[ii]=m["revision"]
            corr=cl["corrections"].get(ii,"")
            corr_map[ii]=corr
            # Guardar en memoria si tiene corrección
            if corr:
                mem_records.append({
                    "item_index":ii,"cluster_id":cl["cluster_id"],
                    "app_name":m["app_name"],"app_address":m["app_address"],
                    "anchor_name":anchor["app_name"] if anchor else "",
                    "correction":corr,"is_new":"NUEVO:" in corr,
                    "file_name":filename,
                })
    if mem_records: mem_save(mem_records)
    # Guardar progreso de revisión
    save_reviewed_clusters(clusters, filename)
    # Guardar resumen del archivo
    ok  = sum(c["ok_count"]  for c in clusters)
    bad = sum(c["bad_count"] for c in clusters)
    save_reviewed_file(filename, session.get("review_type","stores_restaurant"), len(clusters), ok, bad)

    df["item_index"]=df["item_index"].astype(str).str.strip()
    df["revision"]=df["item_index"].map(rev_map).fillna(1).astype(int)
    df["correccion"]=df["item_index"].map(corr_map).fillna("")

    # Hoja 2: correcciones externas acumuladas
    ext_records = ext_corr_list()
    df_ext = pd.DataFrame(ext_records if ext_records else [], columns=[
        "id","store_id","item_index","cluster_index","cluster_name",
        "cluster_address","app_name","app_address","correction","added_at","file_name"
    ])
    # Columnas ordenadas como pidió el usuario
    ext_cols = ["cluster_index","cluster_name","cluster_address",
                "app_name","app_address","item_index","correction"]
    df_ext_out = df_ext[ext_cols] if not df_ext.empty else pd.DataFrame(columns=ext_cols)
    df_ext_out.insert(ext_cols.index("correction")+1, "revision", 0)

    out = UPLOAD_FOLDER/"revisado_final.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Revisión", index=False)
        if not df_ext_out.empty:
            df_ext_out.to_excel(writer, sheet_name="Correcciones externas", index=False)

    # Nombre dinámico: nombre_archivo_revisado_YYYYMMDD_HHMMSS.xlsx
    stem = Path(filename).stem
    ts   = time.strftime("%Y%m%d_%H%M%S")
    download_name = f"{stem}_revisado_{ts}.xlsx"
    return send_file(out, as_attachment=True, download_name=download_name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/dishes/upload", methods=["POST"])
def dishes_upload():
    f = request.files.get("file")
    if not f: return jsonify({"error":"Sin archivo"}), 400
    ext  = Path(f.filename).suffix.lower()
    path = UPLOAD_FOLDER / f"dishes{ext}"
    f.save(path)
    try: df = load_file(path)
    except Exception as e: return jsonify({"error":str(e)}), 400
    groups, meta = build_dish_groups(df, filename=f.filename)
    session["dishes_path"]     = str(path)
    session["dishes_groups"]   = groups
    session["dishes_filename"] = f.filename
    session["review_type"]     = "dishes"
    session["is_reviewed"]     = False
    # Intentar restaurar estado previo
    saved_state, saved_at = load_session_state("dishes_"+f.filename)
    restored = False
    if saved_state:
        rev_map = {s["group_id"]: s for s in saved_state}
        for g in groups:
            saved = rev_map.get(g["group_id"])
            if saved:
                g["revision"]    = saved["revision"]
                g["fusion_with"] = saved.get("fusion_with","")
        restored = True
        meta["saved_at"] = saved_at
    session["dishes_groups"] = groups; session.modified = True
    # Si el archivo ya fue marcado como revisado antes, restaurar ese estado
    already_reviewed = False
    try:
        conn = sqlite3.connect(DB_FILE)
        row = conn.execute(
            "SELECT COUNT(*) FROM reviewed_files WHERE file_name=?", (f.filename,)
        ).fetchone()
        conn.close()
        already_reviewed = row and row[0] > 0
    except: pass
    session["is_reviewed"] = already_reviewed
    session.modified = True
    return jsonify({"groups": groups, "meta": meta,
                    "filename": f.filename, "restored": restored,
                    "already_reviewed": already_reviewed})

@app.route("/dishes/update", methods=["POST"])
def dishes_update():
    data     = request.json
    group_id = data["group_id"]
    revision = data["revision"]
    fusion   = data.get("fusion_with","")
    groups   = session.get("dishes_groups", [])
    for g in groups:
        if g["group_id"] == group_id:
            g["revision"]    = revision
            g["fusion_with"] = fusion
            # Si marca como 0, recalcular categorías de error
            if revision in (0, 2):
                g["error_cats"] = detect_error_categories(g["members"])
            else:
                g["error_cats"] = []
            break
            break
    session["dishes_groups"] = groups; session.modified = True
    state = [{"group_id": g["group_id"], "revision": g["revision"],
              "fusion_with": g.get("fusion_with","")} for g in groups]
    filename = session.get("dishes_filename","")
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT OR REPLACE INTO session_state (file_name, state_json, saved_at)
        VALUES (?,?,?)
    """, ("dishes_"+filename, json.dumps(state), time.time()))
    conn.commit(); conn.close()
    ok  = sum(1 for g in groups if g["revision"]==1)
    bad = sum(1 for g in groups if g["revision"]==0)
    inc = sum(1 for g in groups if g["revision"]==2)
    return jsonify({"ok": ok, "bad": bad, "incomplete": inc})

@app.route("/dishes/set_fusion", methods=["POST"])
def dishes_set_fusion():
    """Marca dos grupos para fusionar (revision=2) con el menor ID como destino."""
    data   = request.json
    gid_a  = data["group_a"]
    gid_b  = data["group_b"]
    groups = session.get("dishes_groups",[])
    # El destino es el menor group_id
    try:
        dest = str(min(int(gid_a), int(gid_b)))
    except:
        dest = min(gid_a, gid_b)
    for g in groups:
        if g["group_id"] in (gid_a, gid_b):
            g["revision"]    = 2
            g["fusion_with"] = dest
            g["error_cats"]  = []
    session["dishes_groups"] = groups; session.modified = True
    # Guardar par etiquetado label=2 (mismo plato, deben fusionarse)
    ga = next((g for g in groups if g["group_id"]==gid_a), None)
    gb = next((g for g in groups if g["group_id"]==gid_b), None)
    if ga and gb and ga.get("members") and gb.get("members"):
        save_dish_pair(
            ga["members"][0].get("product_name",""),
            ga["members"][0].get("product_description",""),
            gb["members"][0].get("product_name",""),
            gb["members"][0].get("product_description",""),
            label=2, source="fusion",
            file_name=session.get("dishes_filename","")
        )
    # Guardar estado
    state = [{"group_id": g["group_id"], "revision": g["revision"],
              "fusion_with": g.get("fusion_with","")} for g in groups]
    filename = session.get("dishes_filename","")
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""INSERT OR REPLACE INTO session_state (file_name, state_json, saved_at)
        VALUES (?,?,?)""", ("dishes_"+filename, json.dumps(state), time.time()))
    conn.commit(); conn.close()
    return jsonify({"ok": True, "dest": dest})

@app.route("/dishes/export")
def dishes_export():
    dishes_path = session.get("dishes_path")
    groups      = session.get("dishes_groups", [])
    filename    = session.get("dishes_filename","archivo")
    if not dishes_path: return jsonify({"error":"Sin sesión"}), 400
    df = load_file(dishes_path)
    rev_map    = {g["group_id"]: g["revision"]           for g in groups}
    fusion_map = {g["group_id"]: g.get("fusion_with","") for g in groups}
    cats_map   = {g["group_id"]: ",".join(g.get("error_cats",[])) for g in groups}
    df["grupo_pn9_max"] = df["grupo_pn9_max"].astype(str).str.strip()
    df["revision"]      = df["grupo_pn9_max"].map(rev_map).fillna(1).astype(int)
    df["fusion"]        = df["grupo_pn9_max"].map(fusion_map).fillna("")
    df["error_cats"]    = df["grupo_pn9_max"].map(cats_map).fillna("")
    # Guardar en memoria
    conn = sqlite3.connect(DB_FILE)
    for g in groups:
        conn.execute("""
            INSERT INTO reviewed_clusters
              (cluster_index, cluster_name, had_errors, corrections_count, reviewed_at, file_name)
            VALUES (?,?,?,?,?,?)
        """, (g["group_id"], "", 1 if g["revision"] in (0,2) else 0,
              0, time.time(), filename))
    conn.commit(); conn.close()
    # Guardar resumen del archivo
    ok  = sum(1 for g in groups if g["revision"]==1)
    bad = sum(1 for g in groups if g["revision"]==0)
    inc = sum(1 for g in groups if g["revision"]==2)
    save_reviewed_file(filename, "dishes", len(groups), ok, bad, inc)
    out = UPLOAD_FOLDER/"dishes_revisado.xlsx"
    df.to_excel(out, index=False)
    stem = Path(filename).stem
    ts   = time.strftime("%Y%m%d_%H%M%S")
    download_name = f"{stem}_revisado_{ts}.xlsx"
    return send_file(out, as_attachment=True,
                     download_name=download_name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/mark_as_reviewed", methods=["POST"])
def mark_as_reviewed():
    """Registra el archivo como revisado en memoria sin descargar."""
    review_type = session.get("review_type","stores_restaurant")
    pairs_added = 0
    if review_type == "dishes":
        groups   = session.get("dishes_groups",[])
        filename = session.get("dishes_filename","")
        ok  = sum(1 for g in groups if g["revision"]==1)
        bad = sum(1 for g in groups if g["revision"]==0)
        inc = sum(1 for g in groups if g["revision"]==2)
        save_reviewed_file(filename, "dishes", len(groups), ok, bad, inc)
        save_reviewed_clusters_dishes(groups, filename)
        session["is_reviewed"] = True
        session.modified = True
    else:
        clusters = session.get("clusters",[])
        filename = session.get("filename","")
        ok  = sum(c["ok_count"]  for c in clusters)
        bad = sum(c["bad_count"] for c in clusters)
        save_reviewed_clusters(clusters, filename)
        save_reviewed_file(filename, review_type, len(clusters), ok, bad)
        # Registrar feedback final
        fb = load_feedback()
        for cl in clusters:
            anchor = next((m for m in cl["members"] if m["is_anchor"]), None)
            if not anchor: continue
            for m in cl["members"]:
                if m["is_anchor"] or m.get("score") is None: continue
                model_rev = 1 if m["score"] >= cl.get("threshold_used", 0.80) else 0
                human_rev = m["revision"]
                if model_rev != human_rev:
                    save_store_pair(
                        anchor["app_name"], anchor["app_address"], cl["cluster_id"],
                        m["app_name"], m["app_address"], cl["cluster_id"],
                        label=human_rev, source="reviewed", file_name=filename
                    )
                    fb["pairs"].append({
                        "anchor_name": anchor["app_name"],
                        "anchor_addr": anchor["app_address"],
                        "member_name": m["app_name"],
                        "member_addr": m["app_address"],
                        "model_score": m["score"],
                        "model_rev":   model_rev,
                        "human_rev":   human_rev,
                        "ts":          time.time()
                    })
                    fb["stats"]["total"] += 1
                    if human_rev == 1: fb["stats"]["overrides_to_1"] += 1
                    else:              fb["stats"]["overrides_to_0"] += 1
                    pairs_added += 1
        recent = fb["pairs"][-30:]
        if len(recent) >= 5:
            fp  = sum(1 for p in recent if p["model_rev"]==1 and p["human_rev"]==0)
            fn  = sum(1 for p in recent if p["model_rev"]==0 and p["human_rev"]==1)
            adj = max(-0.10, min(0.10, (fp-fn)/len(recent)*0.10))
            fb["threshold_adj"] = round(adj, 4)
        save_feedback(fb)
        session["is_reviewed"] = True
        session.modified = True
    return jsonify({"ok": True, "filename": filename, "pairs_added": pairs_added})

@app.route("/unmark_reviewed", methods=["POST"])
def unmark_reviewed():
    """Desmarca el archivo como revisado, permitiendo seguir editando correcciones.
    Borra los registros de reviewed_clusters y reviewed_files para este archivo,
    evitando que queden entradas huérfanas de una revisión incompleta."""
    filename = session.get("filename", "")
    if not filename:
        return jsonify({"error": "Sin archivo activo"}), 400

    conn = sqlite3.connect(DB_FILE)
    conn.execute("DELETE FROM reviewed_clusters WHERE file_name = ?", (filename,))
    conn.execute("DELETE FROM reviewed_files    WHERE file_name = ?", (filename,))
    conn.commit(); conn.close()

    session["is_reviewed"] = False
    session.modified = True
    return jsonify({"ok": True})

@app.route("/reset_session", methods=["POST"])
def reset_session():
    """Limpia la sesión activa para empezar una nueva revisión."""
    for key in ["clusters","review_path","filename","country",
                "dishes_groups","dishes_path","dishes_filename"]:
        session.pop(key, None)
    session.modified = True
    return jsonify({"ok": True})

@app.route("/reset_file", methods=["POST"])
def reset_file():
    """Borra TODO el estado de un archivo: sesión Flask + todas las tablas de memory.db
    relacionadas. Deja el archivo como si nunca hubiera sido abierto."""
    filename = session.get("filename") or (request.json or {}).get("filename", "")
    if not filename:
        return jsonify({"error": "Sin archivo activo"}), 400

    conn = sqlite3.connect(DB_FILE)
    conn.execute("DELETE FROM session_state        WHERE file_name = ?", (filename,))
    conn.execute("DELETE FROM reviewed_clusters    WHERE file_name = ?", (filename,))
    conn.execute("DELETE FROM reviewed_files       WHERE file_name = ?", (filename,))
    conn.execute("DELETE FROM corrections          WHERE file_name = ?", (filename,))
    conn.execute("DELETE FROM external_corrections WHERE file_name = ?", (filename,))
    conn.commit(); conn.close()

    # Conservar review_type para no perder el contexto de tipo de archivo
    rt = session.get("review_type", "stores_restaurant")
    for key in ["clusters", "review_path", "country",
                "dishes_groups", "dishes_path", "dishes_filename", "is_reviewed"]:
        session.pop(key, None)
    session["filename"]    = filename
    session["review_type"] = rt
    session.modified = True

    return jsonify({"ok": True, "filename": filename})

@app.route("/upload_bd", methods=["POST"])
def upload_bd():
    f          = request.files.get("file")
    cluster_id = request.form.get("cluster_id","")
    if not f: return jsonify({"error":"Sin archivo"}), 400
    ext  = Path(f.filename).suffix.lower()
    path = UPLOAD_FOLDER / f"bd_{re.sub(r'[^a-z0-9]','_',cluster_id.lower())}{ext}"
    f.save(path)
    try: df = load_file(path)
    except Exception as e: return jsonify({"error":str(e)}), 400
    from flask import jsonify as _j
    rows = [{k:("" if pd.isna(v) else str(v)) for k,v in row.items()} for _,row in df.iterrows()]
    by_cluster = {}
    for row in rows:
        ci = row.get("cluster_index","")
        by_cluster.setdefault(ci,[]).append(row)
    clusters_found = [{"cluster_index":ci,"members":mems,"is_clean":False} for ci,mems in by_cluster.items()]
    return jsonify({"cluster_id":cluster_id,"clusters_found":clusters_found})


@app.route("/ext_correction/add", methods=["POST"])
def ext_correction_add():
    data = request.json
    data["file_name"] = session.get("filename","")
    # Si scraper_source está vacío, buscarlo en la BD por store_id
    if not data.get("scraper_source","").strip() and data.get("store_id","").strip():
        store_id = data["store_id"].strip()
        table = get_pg_table()
        sql = f"SELECT scraper_source FROM {table} WHERE store_id = '{store_id}' LIMIT 1"
        rows, err = pg_query(sql)
        if not err and rows:
            data["scraper_source"] = rows[0].get("scraper_source","")
    ext_corr_add(data)
    # Guardar par etiquetado: store externo es distinto al cluster donde estaba
    save_store_pair(
        data.get("app_name",""), data.get("app_address",""), data.get("cluster_index",""),
        "", "", data.get("correction",""),
        label=0, source="ext_correction", file_name=data["file_name"]
    )
    return jsonify({"ok": True, "count": len(ext_corr_list())})

@app.route("/ext_correction/list")
def ext_correction_list():
    return jsonify({"records": ext_corr_list()})

@app.route("/ext_correction/delete", methods=["POST"])
def ext_correction_delete():
    ext_corr_delete(request.json.get("id"))
    return jsonify({"ok": True})

def progress_stats_dishes():
    """Progreso de revisión de dishes — solo desde memoria interna."""
    files = get_reviewed_files("dishes")
    total_files = len(files)
    total_groups = sum(f["total"] for f in files)
    total_ok     = sum(f["ok"]    for f in files)
    total_bad    = sum(f["bad"]   for f in files)
    total_inc    = sum(f["incomplete"] for f in files)
    return {
        "mode":          "dishes",
        "total_files":   total_files,
        "total_groups":  total_groups,
        "total_ok":      total_ok,
        "total_bad":     total_bad,
        "total_incomplete": total_inc,
        "files":         files
    }

@app.route("/reviewed_files/clean", methods=["POST"])
def reviewed_files_clean():
    """Elimina duplicados en reviewed_files, dejando solo el más reciente por archivo."""
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        DELETE FROM reviewed_files
        WHERE id NOT IN (
            SELECT MAX(id) FROM reviewed_files GROUP BY file_name, review_type
        )
    """)
    conn.commit()
    remaining = conn.execute("SELECT COUNT(*) FROM reviewed_files").fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "remaining": remaining})

@app.route("/reviewed_files")
def reviewed_files_route():
    review_type = request.args.get("type")
    return jsonify({"files": get_reviewed_files(review_type)})

@app.route("/progress")
def progress():
    mode        = request.args.get("mode", "stores")
    table       = request.args.get("table", "sales_opportunity.dim_maestra")
    country     = request.args.get("country", "mx")
    review_type = request.args.get("review_type", "")

    if mode == "dishes":
        return jsonify(progress_stats_dishes())

    # Determinar tabla y review_type
    if "retail" in table:
        review_type = review_type or "stores_retail"
    else:
        review_type = review_type or "stores_restaurant"

    stats, err = progress_stats(table=table, country=country, review_type=review_type)
    if err: return jsonify({"error": err}), 400
    stats["files"] = get_reviewed_files(review_type)
    return jsonify(stats)

@app.route("/memory/reset", methods=["POST"])
def memory_reset():
    conn=sqlite3.connect(DB_FILE)
    conn.execute("DELETE FROM corrections"); conn.commit(); conn.close()
    return jsonify({"ok":True})


@app.route("/dishes/suggestions", methods=["POST"])
def dishes_suggestions():
    """Retorna los 3 grupos más similares al grupo dado."""
    data     = request.json
    group_id = data["group_id"]
    groups   = session.get("dishes_groups", [])
    target   = next((g for g in groups if g["group_id"]==group_id), None)
    if not target: return jsonify({"suggestions":[]})

    others = [g for g in groups if g["group_id"] != group_id]
    scored = []
    for g in others:
        sim = compute_group_similarity(target, g)
        scored.append({"group_id": g["group_id"], "sim": sim,
                       "members": g["members"], "count": g["count"]})
    scored.sort(key=lambda x: x["sim"], reverse=True)
    return jsonify({"suggestions": scored[:3]})


@app.route("/store_pairs/stats")
def store_pairs_stats():
    """Estadísticas de pares etiquetados de stores."""
    conn = sqlite3.connect(DB_FILE)
    total    = conn.execute("SELECT COUNT(*) FROM store_pairs").fetchone()[0]
    by_src   = conn.execute("""
        SELECT source, COUNT(*) FROM store_pairs GROUP BY source
    """).fetchall()
    conn.close()
    return jsonify({
        "total": total,
        "by_source": {r[0]: r[1] for r in by_src}
    })

@app.route("/save_pair", methods=["POST"])
def save_pair():
    """Guarda un par etiquetado de stores para feedback/fine-tuning.
    Llamado automáticamente al fusionar (label=1) o separar (label=0) subgrupos.
    """
    d = request.json or {}
    file_name = d.get("file_name") or session.get("filename", "")
    save_store_pair(
        app_name_a    = d.get("name_a",""),
        app_address_a = d.get("addr_a",""),
        cluster_index_a = d.get("ci_a",""),
        app_name_b    = d.get("name_b",""),
        app_address_b = d.get("addr_b",""),
        cluster_index_b = d.get("ci_b",""),
        label         = int(d.get("label", 1)),
        source        = d.get("source", "subgroup_drag"),
        file_name     = file_name
    )
    return jsonify({"ok": True})


def clean_correction(corr):
    """Elimina prefijo NUEVO: si existe (artefacto de versiones anteriores)."""
    if isinstance(corr, str) and corr.startswith("NUEVO:"):
        return corr[6:]
    return corr

@app.route("/bd_update/preview", methods=["POST"])
def bd_update_preview():
    """Genera preview enriquecido de los INSERTs que se harían en ctrl_restaurant_homologation.
    No requiere is_reviewed — el preview es de solo lectura.
    El execute sí lo requiere."""

    clusters   = session.get("clusters", [])
    country    = session.get("country", "mx")
    rt         = session.get("review_type", "stores_restaurant")
    store_type = "retail" if rt == "stores_retail" else "restaurant"
    table      = get_pg_table()
    rows = []

    # Correcciones de subgrupos
    for cl in clusters:
        for ii, corr in cl.get("corrections", {}).items():
            if not corr: continue
            corr = clean_correction(corr)
            member = next((m for m in cl["members"] if m["item_index"] == ii), None)
            if not member: continue
            rows.append({
                "store_id":       member.get("store_id", ""),
                "app_name":       member.get("app_name", ""),
                "app_address":    member.get("app_address", ""),
                "scraper_source": member.get("scraper_source", ""),
                "old_cluster":    member.get("cluster_index", cl["cluster_id"]),
                "new_cluster":    corr,
                "country":        country,
                "store_type":     store_type,
                "source":         "file",
            })

    # Correcciones externas (Anotar) — solo del archivo actual
    current_file = session.get("filename", "")
    for rec in ext_corr_list():
        if not rec.get("correction"): continue
        if rec.get("file_name") and rec["file_name"] != current_file: continue
        rows.append({
            "store_id":       rec.get("store_id", ""),
            "app_name":       rec.get("app_name", ""),
            "app_address":    rec.get("app_address", ""),
            "scraper_source": rec.get("scraper_source", ""),
            "old_cluster":    rec.get("cluster_index", ""),
            "new_cluster":    clean_correction(rec.get("correction", "")),
            "country":        country,
            "store_type":     store_type,
            "source":         "external",
        })

    if not rows:
        return jsonify({"rows": [], "total": 0, "store_type": store_type, "country": country})

    # Deduplicar por store_id: si el mismo store fue corregido varias veces, gana la última
    seen_sid = {}
    for r in rows:
        sid = r.get("store_id", "")
        seen_sid[sid if sid else id(r)] = r
    rows = list(seen_sid.values())

    # Enriquecer con contexto de PostgreSQL
    # Traer cluster_name + miembros de old y new clusters únicos
    old_clusters = list({r["old_cluster"] for r in rows if r["old_cluster"]})
    new_clusters = list({r["new_cluster"] for r in rows if r["new_cluster"] and not r["new_cluster"].startswith("NUEVO:")})
    all_clusters = list(set(old_clusters + new_clusters))

    cluster_ctx = {}  # cluster_index → {name, address, members, prev_corrections}
    if all_clusters and pg_is_configured():
        placeholders = ",".join(f"'{c}'" for c in all_clusters)
        sql_ctx = f"""
            SELECT cluster_index, cluster_name, cluster_address,
                   store_id, app_name, app_address, scraper_source, item_index
            FROM {table}
            WHERE cluster_index IN ({placeholders})
            AND country = '{country}'
            ORDER BY cluster_index, item_index
        """
        ctx_rows, err = pg_query(sql_ctx, timeout_ms=15000)
        if not err and ctx_rows:
            for r in ctx_rows:
                ci = r["cluster_index"]
                if ci not in cluster_ctx:
                    cluster_ctx[ci] = {
                        "name":    r["cluster_name"] or "",
                        "address": r["cluster_address"] or "",
                        "members": []
                    }
                cluster_ctx[ci]["members"].append({
                    "store_id":    r["store_id"],
                    "app_name":    r["app_name"],
                    "app_address": r["app_address"],
                    "is_anchor":   r["item_index"] == ci
                })

        # Traer correcciones previas aplicadas a estos clusters
        prev_sql = f"""
            SELECT old_cluster, new_cluster, store_id, load_date
            FROM sales_opportunity.ctrl_restaurant_homologation
            WHERE (old_cluster IN ({placeholders}) OR new_cluster IN ({placeholders}))
            ORDER BY load_date DESC
            LIMIT 100
        """
        prev_rows, _ = pg_query(prev_sql, timeout_ms=10000)
        prev_by_cluster = {}
        if prev_rows:
            for r in prev_rows:
                for ci in [r["old_cluster"], r["new_cluster"]]:
                    if ci in all_clusters:
                        prev_by_cluster.setdefault(ci, []).append({
                            "store_id":    r["store_id"],
                            "old_cluster": r["old_cluster"],
                            "new_cluster": r["new_cluster"],
                            "load_date":   str(r["load_date"])[:10] if r["load_date"] else ""
                        })
        for ci in cluster_ctx:
            cluster_ctx[ci]["prev_corrections"] = prev_by_cluster.get(ci, [])

    # Adjuntar contexto a cada row
    for r in rows:
        r["old_ctx"] = cluster_ctx.get(r["old_cluster"], {})
        r["new_ctx"] = cluster_ctx.get(r["new_cluster"], {})

    return jsonify({"rows": rows, "total": len(rows),
                    "store_type": store_type, "country": country})


@app.route("/bd_update/execute", methods=["POST"])
def bd_update_execute():
    """Ejecuta los INSERTs en ctrl_restaurant_homologation.
    Si hay duplicados por store_id (mismo store corregido dos veces), gana el último."""
    if not pg_is_configured():
        return jsonify({"error": "BD no configurada"}), 400

    rows_raw = request.json.get("rows", [])

    # Deduplicar: por store_id gana la última corrección recibida (last-wins)
    seen_sid = {}
    for r in rows_raw:
        sid = r.get("store_id", "")
        seen_sid[sid if sid else id(r)] = r
    rows = list(seen_sid.values())

    inserted = 0
    errors   = []

    conn_str = (f"host={os.environ.get('PG_HOST')} "
                f"port={os.environ.get('PG_PORT','5432')} "
                f"dbname={os.environ.get('PG_DATABASE')} "
                f"user={os.environ.get('PG_USER')} "
                f"password={os.environ.get('PG_PASSWORD')}")
    try:
        import psycopg2
        conn = psycopg2.connect(conn_str)
        cur  = conn.cursor()
        for r in rows:
            try:
                cur.execute("""
                    INSERT INTO sales_opportunity.ctrl_restaurant_homologation
                      (old_cluster, new_cluster, store_id, scraper_source,
                       country, load_date, homologation_type, store_type, data_quality)
                    VALUES (%s, %s, %s, %s, %s, NOW(), %s, %s, %s)
                """, (
                    r["old_cluster"], r["new_cluster"],
                    r["store_id"],    r.get("scraper_source",""),
                    r["country"],     "manual",
                    r["store_type"],  True
                ))
                inserted += 1
            except Exception as e:
                errors.append(str(e))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True, "inserted": inserted, "errors": errors})


@app.route("/audit_cluster", methods=["POST"])
def audit_cluster():
    """
    Dado un cluster_index, trae todos sus miembros desde dim_maestra,
    calcula scores vs el ancla, y busca stores dispersas (similares al ancla
    en otros clusters).
    """
    if not pg_is_configured():
        return jsonify({"error": "BD no configurada"}), 400

    data          = request.json
    cluster_index = data.get("cluster_index", "").strip()
    country       = data.get("country", session.get("country", "mx"))
    search_name   = data.get("search_name", "").strip()  # store del subgrupo origen
    search_addr   = data.get("search_addr", "").strip()
    table         = get_pg_table()

    if not cluster_index:
        return jsonify({"error": "cluster_index requerido"}), 400

    # 1. Traer todos los miembros del cluster
    sql_members = f"""
        SELECT cluster_index, cluster_name, cluster_address,
               store_id, app_name, app_address, scraper_source, item_index,
               cluster_latitude, cluster_longitude
        FROM {table}
        WHERE cluster_index = '{cluster_index}'
        AND country = '{country}'
        ORDER BY item_index
    """
    members_raw, err = pg_query(sql_members, timeout_ms=10000)
    if err: return jsonify({"error": err}), 400
    if not members_raw: return jsonify({"error": f"Cluster {cluster_index} no encontrado"}), 404

    # Identificar ancla (item_index == cluster_index)
    anchor = next((m for m in members_raw if m["item_index"] == cluster_index), None)
    if not anchor:
        anchor = members_raw[0]  # fallback: primer miembro

    anchor_name = anchor.get("app_name", "")
    anchor_addr = anchor.get("app_address", "")

    # 2. Calcular scores de cada miembro vs el ancla
    member_names  = [m.get("app_name","")    for m in members_raw]
    member_addrs  = [m.get("app_address","") for m in members_raw]
    scores = batch_semantic_sim(anchor_name, anchor_addr,
                                [{"app_name": n, "app_address": a}
                                 for n, a in zip(member_names, member_addrs)])

    members = []
    for i, m in enumerate(members_raw):
        is_anchor = m["item_index"] == cluster_index
        members.append({
            "item_index":    m["item_index"],
            "store_id":      m["store_id"],
            "app_name":      m["app_name"],
            "app_address":   m["app_address"],
            "scraper_source":m.get("scraper_source",""),
            "is_anchor":     is_anchor,
            "score":         1.0 if is_anchor else round(scores[i], 3),
        })

    # 3. Buscar dispersos
    # Si viene search_name/addr del subgrupo, usarlos — encuentran el cluster correcto
    # aunque sea distinto al ancla del cluster BD
    disp_name = search_name if search_name else anchor_name
    disp_addr = search_addr if search_addr else anchor_addr

    disp_bwords = brand_words(disp_name)
    disp_akeys  = extract_addr_keys(disp_addr)

    dispersos = []
    if disp_bwords:
        name_cond = " and ".join(word_ilike("app_name", w) for w in disp_bwords[:2])
        addr_parts = []
        if disp_akeys:
            addr_parts = [addr_ilike_safe("app_address", w) for w in disp_akeys[:2]]

        where_parts = [f"({name_cond})"]
        if addr_parts:
            where_parts.append("(" + " and ".join(addr_parts) + ")")
        where_clause = " and ".join(where_parts)

        sql_dispersos = f"""
            SELECT cluster_index, cluster_name, cluster_address,
                   store_id, app_name, app_address, scraper_source, item_index
            FROM {table}
            WHERE country = '{country}'
            AND cluster_index != '{cluster_index}'
            AND {where_clause}
            ORDER BY cluster_index
            LIMIT 50
        """
        disp_raw, err2 = pg_query(sql_dispersos, timeout_ms=15000)
        if not err2 and disp_raw:
            # Agrupar por cluster
            disp_by_cluster = {}
            for r in disp_raw:
                ci = r["cluster_index"]
                if ci not in disp_by_cluster:
                    disp_by_cluster[ci] = {
                        "cluster_index": ci,
                        "cluster_name":  r.get("cluster_name",""),
                        "cluster_address":r.get("cluster_address",""),
                        "members": []
                    }
                is_anchor_disp = r["item_index"] == ci
                disp_by_cluster[ci]["members"].append({
                    "store_id":      r["store_id"],
                    "app_name":      r["app_name"],
                    "app_address":   r["app_address"],
                    "scraper_source":r.get("scraper_source",""),
                    "item_index":    r["item_index"],
                    "is_anchor":     is_anchor_disp,
                })
            dispersos = list(disp_by_cluster.values())

    # 4. Correcciones previas en este cluster
    prev_sql = f"""
        SELECT store_id, old_cluster, new_cluster, load_date
        FROM sales_opportunity.ctrl_restaurant_homologation
        WHERE old_cluster = '{cluster_index}' OR new_cluster = '{cluster_index}'
        ORDER BY load_date DESC LIMIT 20
    """
    prev_rows, _ = pg_query(prev_sql, timeout_ms=10000)
    prev_corrections = []
    if prev_rows:
        prev_corrections = [{"store_id": r["store_id"],
                             "old_cluster": r["old_cluster"],
                             "new_cluster": r["new_cluster"],
                             "load_date":   str(r["load_date"])[:10] if r["load_date"] else ""}
                            for r in prev_rows]

    return jsonify({
        "cluster_index":    cluster_index,
        "cluster_name":     anchor.get("cluster_name",""),
        "cluster_address":  anchor.get("cluster_address",""),
        "anchor":           anchor_name,
        "anchor_addr":      anchor_addr,
        "members":          members,
        "dispersos":        dispersos,
        "prev_corrections": prev_corrections,
        "threshold":        get_threshold(),
    })

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO PLACES — scraping de malls, dark kitchens, centros comerciales
# ═══════════════════════════════════════════════════════════════════════════════

# Import lazy: scraper sólo se carga si se usa este módulo
def _get_scraper():
    import scraper as _s
    return _s


@app.route("/places/stats")
def places_stats():
    s = _get_scraper()
    return jsonify(s.places_stats())


@app.route("/places/list")
def places_list():
    s       = _get_scraper()
    country = request.args.get("country", "")
    ptype   = request.args.get("place_type", "")
    query   = request.args.get("q", "")
    limit   = int(request.args.get("limit", 500))
    rows    = s.places_list(
        country    = country or None,
        place_type = ptype   or None,
        query      = query   or None,
        limit      = limit,
    )
    return jsonify({"rows": rows, "total": len(rows)})


@app.route("/places/scrape", methods=["POST"])
def places_scrape():
    """
    Lanza un scraping sincrónico.
    Body JSON:
      {
        "source":     "overpass" | "internal_db",
        "place_type": "mall" | "strip_center" | "market" | "commercial" | "dark_kitchen",
        "country":    "cl" | "mx" | "co" | ...,
        "city":       "Santiago"   (opcional, sólo para overpass)
      }
    Retorna progreso y resultado cuando termina.
    """
    s    = _get_scraper()
    data = request.json or {}

    source     = data.get("source", "overpass")
    place_type = data.get("place_type", "mall")
    country    = data.get("country", "cl").lower()
    city       = data.get("city", "").strip() or None

    messages = []
    def log(msg): messages.append(msg)

    if source == "internal_db":
        if not pg_is_configured():
            return jsonify({"error": "BD no configurada — agrega credenciales en .env"}), 400
        result = s.scrape_dark_kitchens_db(
            pg_query_fn  = pg_query,
            country      = country,
            min_brands   = int(data.get("min_brands", 4)),
            min_stores   = int(data.get("min_stores", 5)),
            progress_cb  = log,
        )
    else:
        result = s.scrape_overpass(
            place_type  = place_type,
            country     = country,
            city        = city,
            progress_cb = log,
        )

    result["log"] = messages
    return jsonify(result)


@app.route("/places/delete/<int:place_id>", methods=["DELETE"])
def places_delete(place_id):
    s = _get_scraper()
    s.places_delete(place_id)
    return jsonify({"ok": True, "deleted": place_id})


@app.route("/places/export")
def places_export():
    """Genera y descarga el Excel con filtros opcionales."""
    s       = _get_scraper()
    country = request.args.get("country", "") or None
    ptype   = request.args.get("place_type", "") or None
    query   = request.args.get("q", "") or None

    path, err = s.export_places_excel(country=country, place_type=ptype, query=query)
    if err:
        return jsonify({"error": err}), 400

    from flask import send_file as _sf
    import os
    filename = f"places_{country or 'all'}_{ptype or 'all'}.xlsx"
    return _sf(path, as_attachment=True, download_name=filename,
               mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
@app.route("/places/verify/<int:place_id>", methods=["POST"])
def places_verify(place_id):
    verified = request.json.get("verified", 1)
    conn = sqlite3.connect(DB_FILE)
    conn.execute("UPDATE places SET verified=? WHERE id=?", (verified, place_id))
    conn.commit(); conn.close()
    return jsonify({"ok": True, "id": place_id, "verified": verified})

@app.route("/cluster_members", methods=["POST"])
def cluster_members():
    """Trae todos los members de un cluster_index específico."""
    data        = request.json or {}
    cluster_id  = data.get("cluster_id")
    review_type = data.get("review_type", session.get("review_type", "stores_restaurant"))
    if not cluster_id:
        return jsonify({"error": "cluster_id requerido"})
    table = ("sales_opportunity.dim_maestra_retail"
             if review_type == "stores_retail"
             else "sales_opportunity.dim_maestra")
    sql = f"""
        SELECT cluster_index, store_id, item_index, cluster_name, cluster_address,
               app_name, app_address, app_longitude, app_latitude, scraper_source,
               cluster_latitude, cluster_longitude, cluster_ciudad,
               (item_index = cluster_index) AS is_anchor
        FROM {table}
        WHERE cluster_index = %s
        ORDER BY is_anchor DESC, store_id
    """
    try:
        import psycopg2, psycopg2.extras
        cfg = get_pg_config()
        conn = psycopg2.connect(**cfg, connect_timeout=15)
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, (cluster_id,))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        if not rows:
            return jsonify({"error": "sin datos", "cluster_id": cluster_id})
        members = [{
            "item_index":    str(r["item_index"]),
            "store_id":      str(r["store_id"]),
            "cluster_index": str(r["cluster_index"]),
            "app_name":      str(r["app_name"] or ""),
            "app_address":   str(r["app_address"] or ""),
            "app_latitude":  float(r["app_latitude"]) if r["app_latitude"] else None,
            "app_longitude": float(r["app_longitude"]) if r["app_longitude"] else None,
            "scraper_source": str(r["scraper_source"] or ""),
            "is_anchor":     bool(r["is_anchor"]),
            "revision":      0, "correccion": "",
        } for r in rows]
        anchor = next((m for m in members if m["is_anchor"]), members[0])
        return jsonify({
            "cluster_id":     cluster_id,
            "anchor_name":    rows[0].get("cluster_name") or anchor["app_name"],
            "anchor_address": rows[0].get("cluster_address") or anchor["app_address"],
            "members":        members
        })
    except Exception as e:
        return jsonify({"error": str(e)})



def scoring_run():
    import threading
    data        = request.json or {}
    country     = data.get("country", "cl")
    city        = data.get("city")
    review_type = data.get("review_type", "stores_restaurant")
    dist_m      = float(data.get("dist_m", 50))
    threading.Thread(target=run_scoring_job,
                     args=(country, city, review_type, dist_m),
                     daemon=True).start()
    return jsonify({"ok": True, "msg": f"Job iniciado para {country}/{city or 'all'}"})


@app.route("/scoring/status", methods=["GET"])
def scoring_status():
    country     = request.args.get("country", "cl")
    review_type = request.args.get("review_type", "stores_restaurant")
    sc = sqlite3.connect(DB_FILE)
    rows = sc.execute("""
        SELECT cluster_ciudad, main_chain IS NOT NULL AND main_chain != '' AS is_chain,
               estado_revision, COUNT(*) AS cnt,
               AVG(score_t1) AS avg_t1, SUM(has_t2) AS sum_t2
        FROM cluster_scores
        WHERE country = ? AND review_type = ?
        GROUP BY cluster_ciudad, is_chain, estado_revision
        ORDER BY cnt DESC
    """, (country, review_type)).fetchall()
    sc.close()
    return jsonify({"rows": [
        {"ciudad": r[0], "is_chain": bool(r[1]), "estado": r[2],
         "count": r[3], "avg_t1": round(r[4] or 1, 3), "t2_count": r[5]}
        for r in rows
    ]})


@app.route("/scoring/clusters", methods=["POST"])
def scoring_clusters():
    data        = request.json or {}
    country     = data.get("country", "cl")
    review_type = data.get("review_type", "stores_restaurant")
    city        = data.get("city")           # legacy single value
    cities      = data.get("cities")         # list
    estado_pg   = data.get("estado_pg")      # cluster_estado filter
    estados     = data.get("estados")        # list
    chain       = data.get("main_chain")     # legacy single value
    chains      = data.get("chains")         # list
    alert_t1    = data.get("alert_t1")
    alert_t2    = data.get("alert_t2")
    t1_thr      = float(data.get("t1_threshold", 0.70))
    member_min  = data.get("member_min")
    member_max  = data.get("member_max")
    limit       = int(data.get("limit", 100))
    offset      = int(data.get("offset", 0))

    is_chain    = data.get("is_chain")
    estado      = data.get("estado_revision")

    conds  = ["country = ?", "review_type = ?"]
    params = [country, review_type]

    # Ciudad — array o valor único
    all_cities = list(cities or ([city] if city else []))
    if all_cities:
        ph = ",".join(["?"]*len(all_cities))
        conds.append(f"cluster_ciudad IN ({ph})")
        params.extend(all_cities)

    # Estado/región — array
    all_estados = list(estados or [])
    if all_estados:
        ph = ",".join(["?"]*len(all_estados))
        conds.append(f"cluster_estado IN ({ph})")
        params.extend(all_estados)

    # Cadena — array o valor único
    all_chains = list(chains or ([chain] if chain else []))
    if all_chains:
        ph = ",".join(["?"]*len(all_chains))
        conds.append(f"main_chain IN ({ph})")
        params.extend(all_chains)

    if estado:
        conds.append("estado_revision = ?"); params.append(estado)
    if is_chain is True:
        conds.append("main_chain IS NOT NULL AND main_chain != ''")
    elif is_chain is False:
        conds.append("(main_chain IS NULL OR main_chain = '')")
    if alert_t1:
        conds.append("score_t1 < ?"); params.append(t1_thr)
    if alert_t2:
        conds.append("has_t2 = 1")
    if member_min is not None:
        conds.append("member_count >= ?"); params.append(int(member_min))
    if member_max is not None:
        conds.append("member_count <= ?"); params.append(int(member_max))

    where = " AND ".join(conds)
    sc = sqlite3.connect(DB_FILE)
    sc.row_factory = sqlite3.Row
    total = sc.execute(f"SELECT COUNT(*) FROM cluster_scores WHERE {where}",
                       params).fetchone()[0]
    rows = sc.execute(f"""
        SELECT * FROM cluster_scores WHERE {where}
        ORDER BY has_t2 DESC, score_t1 ASC
        LIMIT ? OFFSET ?
    """, params + [limit, offset]).fetchall()
    stats = sc.execute(f"""
        SELECT COUNT(*),
               SUM(CASE WHEN score_t1 < ? THEN 1 ELSE 0 END),
               SUM(has_t2),
               SUM(CASE WHEN estado_revision='pendiente' THEN 1 ELSE 0 END),
               SUM(CASE WHEN estado_revision='ok' THEN 1 ELSE 0 END),
               SUM(CASE WHEN estado_revision='corregido' THEN 1 ELSE 0 END)
        FROM cluster_scores WHERE {where}
    """, [t1_thr] + params).fetchone()
    sc.close()
    return jsonify({
        "total": total,
        "clusters": [dict(r) for r in rows],
        "stats": {"total": stats[0], "t1_alerts": stats[1], "t2_alerts": stats[2],
                  "pendientes": stats[3], "ok": stats[4], "corregidos": stats[5]}
    })


@app.route("/scoring/mark", methods=["POST"])
def scoring_mark():
    data   = request.json or {}
    cid    = data.get("cluster_index")
    estado = data.get("estado_revision", "ok")
    rt     = data.get("review_type", "stores_restaurant")
    if not cid: return jsonify({"error": "cluster_index requerido"})
    sc = sqlite3.connect(DB_FILE)
    sc.execute("UPDATE cluster_scores SET estado_revision=? WHERE cluster_index=? AND review_type=?",
               (estado, cid, rt))
    sc.commit(); sc.close()
    return jsonify({"ok": True})


@app.route("/scoring/filter_options", methods=["GET"])
def scoring_filter_options():
    country = request.args.get("country", "cl")
    rt      = request.args.get("review_type", "stores_restaurant")

    # Intentar desde cluster_scores (rápido, ya calculado)
    sc = sqlite3.connect(DB_FILE)
    ciudades = [r[0] for r in sc.execute(
        "SELECT DISTINCT cluster_ciudad FROM cluster_scores WHERE country=? AND review_type=? AND cluster_ciudad IS NOT NULL ORDER BY cluster_ciudad",
        (country, rt)).fetchall()]
    estados = [r[0] for r in sc.execute(
        "SELECT DISTINCT cluster_estado FROM cluster_scores WHERE country=? AND review_type=? AND cluster_estado IS NOT NULL AND cluster_estado!='' ORDER BY cluster_estado",
        (country, rt)).fetchall()]
    cadenas = [r[0] for r in sc.execute(
        "SELECT DISTINCT main_chain FROM cluster_scores WHERE country=? AND review_type=? AND main_chain IS NOT NULL AND main_chain!='' ORDER BY main_chain",
        (country, rt)).fetchall()]
    sc.close()

    # Fallback a dim_maestra si cluster_scores está vacío
    if not ciudades:
        try:
            import psycopg2, psycopg2.extras
            cfg = get_pg_config()
            if cfg:
                table = ("sales_opportunity.dim_maestra_retail"
                         if rt == "stores_retail"
                         else "sales_opportunity.dim_maestra")
                conn = psycopg2.connect(**cfg, connect_timeout=15)
                cur  = conn.cursor()
                cur.execute(f"""
                    SELECT DISTINCT cluster_ciudad
                    FROM {table} WHERE country=%s AND cluster_ciudad IS NOT NULL
                    ORDER BY cluster_ciudad
                """, (country,))
                ciudades = [r[0] for r in cur.fetchall()]
                cur.execute(f"""
                    SELECT DISTINCT cluster_estado
                    FROM {table} WHERE country=%s AND cluster_estado IS NOT NULL AND cluster_estado!=''
                    ORDER BY cluster_estado
                """, (country,))
                estados = [r[0] for r in cur.fetchall()]
                cur.execute(f"""
                    SELECT DISTINCT main_chain
                    FROM {table} WHERE country=%s AND main_chain IS NOT NULL AND main_chain!=''
                    ORDER BY main_chain
                """, (country,))
                cadenas = [r[0] for r in cur.fetchall()]
                conn.close()
        except Exception as e:
            print(f"[filter_options] fallback PG error: {e}")

    return jsonify({"ciudades": ciudades, "estados": estados, "cadenas": cadenas})


if __name__=="__main__":
    print("\n"+"="*54)
    print("  DQ Matching Tool")
    print(f"  PostgreSQL: {'✓ configurado' if pg_is_configured() else '✗ falta .env'}")
    print(f"  Memoria:    {DB_FILE}")
    print("  http://localhost:5000")
    print("="*54+"\n")
    import threading
    threading.Thread(target=get_model,daemon=True).start()
    app.run(debug=False,port=5000)
